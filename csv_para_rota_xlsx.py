"""
csv_para_rota_xlsx.py

Converte um CSV de entrega no formato "delivery_list" (colunas:
Address, City, State, Contact — sem número de sequência) em um
.xlsx pronto pra subir no Rota Manager (Sequence, Address, City,
State, Address (Original), Contact, Lat, Lon, Geo Fonte).

ESTRATÉGIA DE GEOCODING — cascata, não "ou um ou outro"
=========================================================
Da mais precisa pra mais segura. Cada nível só é tentado se o
anterior não devolver nada:

  1. here_cep            -> HERE Geocoding API com o CEP estruturado
                             (qq=postalCode=...;country=BRA). Motor
                             PRIMÁRIO quando HERE_API_KEY está
                             configurada: não depende de rua/bairro
                             corretos, só do CEP em si.
  2. here_endereco        -> HERE com endereço livre completo (rua +
                             número + bairro + cidade), se a busca
                             estruturada por CEP não trouxer nada.
  3. Cascata Photon (rua/número/bairro vêm do ViaCEP + número extraído
     do próprio CSV), tentando do mais específico pro mais genérico:
       3a. photon_completo    -> rua + número + bairro + cidade
       3b. photon_sem_bairro  -> rua + número + cidade (sem bairro)
       3c. photon_sem_numero  -> rua + bairro + cidade (sem número)
       3d. photon_bairro      -> bairro + cidade (só aproximação)
  4. Só se TODA a cascata Photon falhar, repete a mesma cascata no
     Nominatim (motor secundário — evita bater sem necessidade no
     limite de 1 req/seg do serviço público):
       4a. nominatim_completo
       4b. nominatim_sem_bairro
       4c. nominatim_sem_numero
       4d. nominatim_bairro

Por que a cascata tira o NÚMERO antes do BAIRRO (e nunca o contrário):
em Goiânia é comum o mesmo nome/código de rua existir em setores
diferentes (ex.: "Rua 1025" existe no Setor Pedro Ludovico, mas nomes
parecidos aparecem em outros setores). Se a gente tenta "rua + cidade"
sem bairro, o geocoder pode casar com a rua homônima do lado errado da
cidade — um erro silencioso, pior que não ter coordenada nenhuma,
porque parece certo e não é. Perder o número da casa é uma perda de
precisão aceitável (o pino cai no meio da rua certa); perder o bairro
é uma aposta que pode acertar o setor errado.

Photon (Komoot, https://photon.komoot.io) é a camada gratuita
PRIMÁRIA da cascata OSM porque tolera melhor endereço incompleto ou
com erro de digitação e responde mais rápido por chamada que o
Nominatim. O Nominatim entra só como segunda tentativa, com o mesmo
esqueleto de queries, pra não desperdiçar chamadas no serviço com
limite de 1 req/seg quando o Photon já teria resolvido.

Cada coordenada gerada carrega a "fonte" (qual nível da cascata
funcionou), gravada no cache e exibida no xlsx na coluna "Geo Fonte".

O que o script faz:
  1. Lê o CSV.
  2. Extrai o endereço detalhado que vem entre parênteses no campo
     "Address" (ex.: "(Avenida Circular, 1117, Q58 L12 Ed D.Thiago
     Ap903)").
  3. Extrai o CEP (formato 00000-000 ou 00000000) e o número do
     logradouro (do prefixo do campo Address ou, na falta, do texto
     entre parênteses) de dentro do mesmo campo "Address".
  4. Geocodifica cada par único (CEP, número) via a cascata acima,
     com cache em disco pra não repetir chamadas em execuções
     futuras.
  5. Inventa uma sequência (1, 2, 3, ...) na ordem do CSV.
  6. Gera um .xlsx com as colunas: Sequence | Address | City | State
     | Address (Original) | Contact | Lat | Lon | Geo Fonte | Link Mapa

Uso:
    python csv_para_rota_xlsx.py entrada.csv
    python csv_para_rota_xlsx.py entrada.csv saida.xlsx
    python csv_para_rota_xlsx.py entrada.csv saida.xlsx --no-geo   (pula geocodificação)

Cache: gera/usa um arquivo "cep_cache.json" na mesma pasta do
script, com o mapeamento "CEP|NUMERO" -> {"lat":..,"lon":..,"fonte":..}.
Entradas em formato antigo (só [lat, lon], ou chave sem "|", sem
"fonte") são tratadas como inválidas e re-geocodificadas
automaticamente.
"""

import csv
import json
import os
import re
import sys
import time
from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

try:
    from dotenv import load_dotenv
    load_dotenv()  # lê HERE_API_KEY de um arquivo .env na mesma pasta, se existir
except ImportError:
    pass  # sem python-dotenv instalado: só funciona com variável de ambiente já exportada

_RE_PARENTESES = re.compile(r"\((.*?)\)")
_RE_CEP = re.compile(r"\b(\d{5})-?(\d{3})\b")

_CACHE_PATH = Path(__file__).with_name("cep_cache.json")
_NOMINATIM_URL = "https://nominatim.openstreetmap.org/search"
_PHOTON_URL = "https://photon.komoot.io/api/"
_VIACEP_URL = "https://viacep.com.br/ws/{cep}/json/"
_HERE_GEOCODE_URL = "https://geocode.search.hereapi.com/v1/geocode"
_HEADERS = {"User-Agent": "RotaManager-Geocoder/2.0 (uso interno, contato: moises)"}

# Mesma chave usada pelo Rota Manager. Se não estiver configurada, o script
# pula direto pra cascata Photon -> Nominatim (sem quebrar a execução).
_HERE_API_KEY = os.environ.get("HERE_API_KEY", "").strip()

# Viés geográfico pra Photon/Nominatim priorizarem resultados perto de
# Goiânia quando o nome da rua/bairro for ambíguo. Ajuste se a operação
# for em outra cidade.
_BIAS_LAT = float(os.environ.get("GEO_BIAS_LAT", "-16.6869"))
_BIAS_LON = float(os.environ.get("GEO_BIAS_LON", "-49.2648"))

# Caixa geográfica de sanidade: qualquer coordenada fora desses limites é
# descartada (tratada como falha, cascata continua pro próximo nível/motor),
# mesmo que a fonte tenha "respondido com sucesso". Isso existe porque a
# HERE (e outros geocoders) já devolveram, na prática, coordenada boa mas
# em cidade/país errado pra um CEP específico (ex.: caiu na Arábia Saudita
# pra um CEP de Goiânia) — sem essa checagem, o app aceitaria isso como
# válido e mandaria o entregador pro lugar errado sem nenhum aviso. A
# margem por padrão cobre a região metropolitana de Goiânia com folga;
# ajuste via env se a operação for em outra cidade/UF.
_BBOX_LAT_MIN = float(os.environ.get("GEO_BBOX_LAT_MIN", "-18.5"))
_BBOX_LAT_MAX = float(os.environ.get("GEO_BBOX_LAT_MAX", "-14.0"))
_BBOX_LON_MIN = float(os.environ.get("GEO_BBOX_LON_MIN", "-50.5"))
_BBOX_LON_MAX = float(os.environ.get("GEO_BBOX_LON_MAX", "-47.5"))


def _dentro_da_regiao(lat: float, lon: float) -> bool:
    return _BBOX_LAT_MIN <= lat <= _BBOX_LAT_MAX and _BBOX_LON_MIN <= lon <= _BBOX_LON_MAX

_ultimo_nominatim = [0.0]  # timestamp da última chamada, pra respeitar 1 req/seg
_ultimo_here = [0.0]       # throttle leve pra HERE (bem mais tolerante que o Nominatim)
_ultimo_photon = [0.0]     # throttle leve pra Photon (instância pública, uso justo)

# GEO_DEBUG=1 imprime o motivo real de cada falha (em vez de só devolver None
# silenciosamente), útil pra diagnosticar chave inválida, bloqueio de rede,
# rate limit, etc.
_DEBUG = os.environ.get("GEO_DEBUG", "").strip() == "1"


def _debug(origem: str, msg: str) -> None:
    if _DEBUG:
        print(f"    [debug/{origem}] {msg}")

_FONTE_LABELS = {
    "here_cep": "HERE (CEP exato)",
    "here_endereco": "HERE (endereço completo)",
    "photon_completo": "Photon (rua+número+bairro)",
    "photon_sem_bairro": "Photon (rua+número)",
    "photon_sem_numero": "Photon (rua+bairro, sem número)",
    "photon_bairro": "Photon (aproximado por bairro)",
    "nominatim_completo": "Nominatim (rua+número+bairro)",
    "nominatim_sem_bairro": "Nominatim (rua+número)",
    "nominatim_sem_numero": "Nominatim (rua+bairro, sem número)",
    "nominatim_bairro": "Nominatim (aproximado por bairro)",
}


def _link_maps(lat: float, lon: float) -> str:
    """Monta um link do Google Maps a partir de lat/lon decimais (com sinal),
    já no formato que o Maps aceita direto — evita erro de conversão manual
    pra DMS (N/S/E/W), que pode jogar o pino no hemisfério errado."""
    return f"https://www.google.com/maps?q={lat},{lon}"


def _extrair_endereco_detalhado(address_bruto: str) -> str:
    """Pega o texto dentro dos parênteses (rua + número + complemento).
    Se não encontrar parênteses, devolve o endereço original inteiro."""
    m = _RE_PARENTESES.search(address_bruto or "")
    if m and m.group(1).strip():
        return m.group(1).strip()
    return (address_bruto or "").strip()


def _extrair_endereco_combinado(address_bruto: str) -> str:
    """Junta o prefixo do campo Address (rua, número, cidade, UF — tudo
    antes do CEP) com o endereço detalhado que vem entre parênteses,
    no formato "prefixo, (detalhado)".

    Motivo: às vezes o prefixo tem um dado que o detalhado não tem (ex.:
    o número "511" solto antes do CEP) e às vezes é o contrário (o
    detalhado tem apto/complemento que o prefixo não tem). Juntando os
    dois, nenhuma informação se perde. O CEP em si é removido daqui
    porque já vira colunas/lat-lon separadas.

    Ex.: "R S 4, 511, Goiânia, GO, 74823-450, (Rua S 4 Apto.: 502), 74823450"
      -> "R S 4, 511, Goiânia, GO, (Rua S 4 Apto.: 502)"
    """
    address_bruto = address_bruto or ""
    detalhado = _extrair_endereco_detalhado(address_bruto)
    tem_parenteses = bool(_RE_PARENTESES.search(address_bruto))

    m_cep = _RE_CEP.search(address_bruto)
    if m_cep:
        prefixo = address_bruto[: m_cep.start()]
    elif tem_parenteses:
        # sem CEP no meio: tira o trecho entre parênteses do prefixo,
        # senão o detalhado apareceria duplicado
        prefixo = _RE_PARENTESES.sub("", address_bruto)
    else:
        prefixo = address_bruto

    prefixo = prefixo.strip().rstrip(",").strip()

    if prefixo and tem_parenteses:
        return f"{prefixo}, ({detalhado})"
    if prefixo:
        return prefixo
    return f"({detalhado})" if detalhado else ""


def _extrair_cep(address_bruto: str) -> str | None:
    """Pega o primeiro CEP encontrado no campo Address (00000-000 ou 00000000)."""
    m = _RE_CEP.search(address_bruto or "")
    if not m:
        return None
    return f"{m.group(1)}-{m.group(2)}"


def _numero_de_trecho(trecho: str) -> str | None:
    """Dado um trecho tipo 'Rua X, 443, ...' ou 'Rua X, SN, ...', devolve
    o primeiro número encontrado logo após o nome da rua (segundo campo
    separado por vírgula), ou None se esse campo não tiver dígito nenhum
    (ex.: 'SN', 'S/N', ou o campo seguinte já é a cidade)."""
    tokens = [t.strip() for t in (trecho or "").split(",") if t.strip()]
    if len(tokens) < 2:
        return None
    m = re.search(r"\d+", tokens[1])
    return m.group(0) if m else None


def _extrair_numero(address_bruto: str) -> str | None:
    """Extrai o número do logradouro. Tenta primeiro o prefixo do campo
    Address (antes do CEP, ex.: 'R 1025, 443, Goiânia, GO'); se não achar,
    tenta o texto entre parênteses (ex.: '(Avenida Circular, 1117, Q58...)'),
    já que às vezes o número só aparece ali."""
    address_bruto = address_bruto or ""
    m_cep = _RE_CEP.search(address_bruto)
    prefixo = address_bruto[: m_cep.start()] if m_cep else ""
    numero = _numero_de_trecho(prefixo)
    if numero:
        return numero
    detalhado = _extrair_endereco_detalhado(address_bruto)
    return _numero_de_trecho(detalhado)


def _carregar_cache() -> dict:
    if _CACHE_PATH.exists():
        try:
            return json.loads(_CACHE_PATH.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            return {}
    return {}


def _salvar_cache(cache: dict) -> None:
    try:
        _CACHE_PATH.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")
    except OSError:
        pass


def _chave_cache(cep: str, numero: str | None) -> str:
    return f"{cep}|{numero or ''}"


def _cache_para_coord(entrada) -> tuple | None:
    """Converte uma entrada do cache em (lat, lon, fonte), ou None se
    inválida/ausente. Entradas em formato antigo (lista [lat, lon], sem
    "fonte") são tratadas como inválidas de propósito, pra forçar
    re-geocodificação: não dá pra saber se vieram de uma estratégia
    antiga e menos segura."""
    if entrada is None:
        return None
    if isinstance(entrada, dict) and "fonte" in entrada:
        lat, lon, fonte = entrada.get("lat"), entrada.get("lon"), entrada.get("fonte")
        if lat is None or lon is None:
            return None
        return (lat, lon, fonte)
    return None  # formato antigo -> inválido, será re-geocodificado


def _here_chamar(params: dict) -> tuple | None:
    """Faz UMA chamada à HERE Geocoding API com os params já montados
    (qq= pra busca estruturada ou q= pra busca livre) e devolve
    (lat, lon) do primeiro resultado, ou None se não achar/der erro."""
    espera = 0.15 - (time.monotonic() - _ultimo_here[0])
    if espera > 0:
        time.sleep(espera)
    try:
        resp = requests.get(_HERE_GEOCODE_URL, params=params, headers=_HEADERS, timeout=10)
        _ultimo_here[0] = time.monotonic()
        if resp.status_code != 200:
            _debug("here", f"status {resp.status_code} | params={params} | corpo={resp.text[:200]!r}")
        resp.raise_for_status()
        dados = resp.json()
        items = dados.get("items") or []
        if items:
            pos = items[0].get("position") or {}
            lat, lon = pos.get("lat"), pos.get("lng")
            if lat is not None and lon is not None:
                lat, lon = float(lat), float(lon)
                if _dentro_da_regiao(lat, lon):
                    return lat, lon
                _debug("here", f"descartado por estar fora da região esperada: ({lat}, {lon}) | params={params}")
        else:
            _debug("here", f"0 items pra params={params}")
    except (requests.RequestException, ValueError, KeyError, IndexError) as e:
        _ultimo_here[0] = time.monotonic()
        _debug("here", f"exceção {type(e).__name__}: {e}")
    return None


def _here_buscar_cep(cep_digits: str) -> tuple | None:
    """Busca via HERE usando o CEP estruturado (qq=postalCode=...;country=BRA).
    É o motor PRIMÁRIO quando há chave: não depende de rua/bairro corretos,
    só do CEP — evita o problema de ruas homônimas em setores diferentes."""
    if not _HERE_API_KEY:
        return None
    cep_fmt = f"{cep_digits[:5]}-{cep_digits[5:]}" if len(cep_digits) == 8 else cep_digits
    params = {
        "qq": f"postalCode={cep_fmt};country=BRA",
        "apiKey": _HERE_API_KEY,
        "limit": 1,
    }
    return _here_chamar(params)


def _here_buscar_endereco(query: str) -> tuple | None:
    """Busca via HERE com endereço livre (rua + número + bairro + cidade),
    usada quando a busca estruturada por CEP não retorna nada."""
    if not _HERE_API_KEY:
        return None
    params = {
        "q": query,
        "apiKey": _HERE_API_KEY,
        "limit": 1,
        "in": "countryCode:BRA",
    }
    return _here_chamar(params)


def _photon_buscar(query: str) -> tuple | None:
    """Faz UMA busca no Photon (instância pública do Komoot), com viés
    geográfico pra região configurada e throttle leve (uso justo de um
    serviço público e gratuito)."""
    espera = 0.5 - (time.monotonic() - _ultimo_photon[0])
    if espera > 0:
        time.sleep(espera)
    params = {
        "q": query,
        "limit": 1,
        "lat": _BIAS_LAT,
        "lon": _BIAS_LON,
    }
    try:
        resp = requests.get(_PHOTON_URL, params=params, headers=_HEADERS, timeout=10)
        _ultimo_photon[0] = time.monotonic()
        if resp.status_code != 200:
            _debug("photon", f"status {resp.status_code} | q={query!r} | corpo={resp.text[:200]!r}")
        resp.raise_for_status()
        dados = resp.json()
        features = dados.get("features") or []
        if features:
            coords = (features[0].get("geometry") or {}).get("coordinates")
            if coords and len(coords) == 2:
                lon, lat = coords[0], coords[1]  # GeoJSON: [lon, lat]
                lat, lon = float(lat), float(lon)
                if _dentro_da_regiao(lat, lon):
                    return lat, lon
                _debug("photon", f"descartado por estar fora da região esperada: ({lat}, {lon}) | q={query!r}")
        else:
            _debug("photon", f"0 features pra q={query!r}")
    except (requests.RequestException, ValueError, KeyError, IndexError) as e:
        _ultimo_photon[0] = time.monotonic()
        _debug("photon", f"exceção {type(e).__name__}: {e}")
    return None


def _nominatim_buscar(query: str) -> tuple | None:
    """Faz UMA busca no Nominatim, respeitando o limite de 1 req/seg
    (independente de quantas vezes essa função for chamada em sequência)."""
    espera = 1.1 - (time.monotonic() - _ultimo_nominatim[0])
    if espera > 0:
        time.sleep(espera)
    params = {"q": query, "format": "json", "limit": 1, "countrycodes": "br"}
    try:
        resp = requests.get(_NOMINATIM_URL, params=params, headers=_HEADERS, timeout=10)
        _ultimo_nominatim[0] = time.monotonic()
        if resp.status_code != 200:
            _debug("nominatim", f"status {resp.status_code} | q={query!r} | corpo={resp.text[:200]!r}")
        resp.raise_for_status()
        dados = resp.json()
        if dados:
            lat, lon = float(dados[0]["lat"]), float(dados[0]["lon"])
            if _dentro_da_regiao(lat, lon):
                return lat, lon
            _debug("nominatim", f"descartado por estar fora da região esperada: ({lat}, {lon}) | q={query!r}")
        else:
            _debug("nominatim", f"0 resultados pra q={query!r}")
    except (requests.RequestException, ValueError, KeyError, IndexError) as e:
        _ultimo_nominatim[0] = time.monotonic()
        _debug("nominatim", f"exceção {type(e).__name__}: {e}")
    return None


def _consultar_viacep(cep_digits: str) -> dict | None:
    """Consulta a base oficial de CEPs (ViaCEP) e devolve
    logradouro/bairro/cidade/uf, ou None se o CEP não existir."""
    url = _VIACEP_URL.format(cep=cep_digits)
    try:
        resp = requests.get(url, headers=_HEADERS, timeout=10)
        resp.raise_for_status()
        dados = resp.json()
        if dados.get("erro"):
            return None
        return dados
    except (requests.RequestException, ValueError, KeyError):
        return None


def _montar_niveis_cascata(rua: str, numero: str | None, bairro: str, cidade: str, estado: str) -> list:
    """Monta a lista de queries da cascata, da mais específica pra mais
    genérica, sem repetir queries idênticas (ex.: quando não há número
    ou não há bairro, alguns níveis colapsam no mesmo texto)."""
    niveis = []
    vistos = set()

    def add(nome: str, partes: list):
        query = ", ".join(p for p in partes if p)
        if query and query not in vistos:
            vistos.add(query)
            niveis.append((nome, query))

    if rua and numero:
        add("completo", [f"{rua}, {numero}", bairro, cidade, estado, "Brasil"])
        add("sem_bairro", [f"{rua}, {numero}", cidade, estado, "Brasil"])
    if rua and bairro:
        add("sem_numero", [rua, bairro, cidade, estado, "Brasil"])
    elif rua:
        add("sem_numero", [rua, cidade, estado, "Brasil"])
    if bairro:
        add("bairro", [bairro, cidade, estado, "Brasil"])

    return niveis


def _rodar_cascata(motor, niveis: list, prefixo_fonte: str) -> tuple | None:
    """Tenta cada nível da cascata nesse motor, na ordem, e para no
    primeiro que retornar coordenada."""
    for nome_nivel, query in niveis:
        coord = motor(query)
        if coord:
            return (*coord, f"{prefixo_fonte}_{nome_nivel}")
    return None


def _geocodificar_par(cep: str, numero: str | None, cidade: str = "Goiânia", estado: str = "GO") -> tuple | None:
    """Geocodifica um par (CEP, número) seguindo a cascata completa:
      1) HERE com o CEP estruturado (só se HERE_API_KEY configurada).
      2) HERE com endereço livre completo (rua+número+bairro+cidade).
      3) Cascata Photon: rua+número+bairro -> rua+número -> rua+bairro
         (sem número) -> bairro (aproximação).
      4) A MESMA cascata no Nominatim, só se o Photon inteiro falhar.
    Retorna (lat, lon, fonte) ou None se nada funcionar."""

    cep_digits = cep.replace("-", "")

    # tentativa 1: HERE com o CEP estruturado
    coord = _here_buscar_cep(cep_digits)
    if coord:
        return (*coord, "here_cep")

    # ViaCEP: base oficial CEP -> endereço, usada pelas tentativas seguintes
    info = _consultar_viacep(cep_digits)
    logradouro = bairro = ""
    cidade_via, estado_via = cidade, estado
    if info:
        logradouro = (info.get("logradouro") or "").strip()
        bairro = (info.get("bairro") or "").strip()
        cidade_via = (info.get("localidade") or cidade).strip()
        estado_via = (info.get("uf") or estado).strip()
        _debug("viacep", f"{cep} -> rua={logradouro!r} bairro={bairro!r} cidade={cidade_via!r}")
    else:
        _debug("viacep", f"{cep} -> sem resposta/CEP não encontrado")

    # tentativa 2: HERE com endereço completo (rua + número + bairro)
    if logradouro:
        partes = [f"{logradouro}, {numero}" if numero else logradouro, bairro, cidade_via, estado_via, "Brasil"]
        query = ", ".join(p for p in partes if p)
        coord = _here_buscar_endereco(query)
        if coord:
            return (*coord, "here_endereco")

    # tentativas 3 e 4: cascata Photon, depois a mesma cascata no Nominatim
    niveis = _montar_niveis_cascata(logradouro, numero, bairro, cidade_via, estado_via)
    if niveis:
        resultado = _rodar_cascata(_photon_buscar, niveis, "photon")
        if resultado:
            return resultado
        resultado = _rodar_cascata(_nominatim_buscar, niveis, "nominatim")
        if resultado:
            return resultado

    return None


def _geocodificar_enderecos(pares: set, cidade: str, estado: str) -> dict:
    """Geocodifica um conjunto de pares únicos (cep, numero), usando cache
    em disco. Retorna dict: (cep, numero) -> (lat, lon, fonte) ou None."""
    cache_bruto = _carregar_cache()
    resultado = {}

    novos = [p for p in pares if _cache_para_coord(cache_bruto.get(_chave_cache(*p))) is None]

    if novos:
        motor = "HERE + cascata Photon/Nominatim" if _HERE_API_KEY else "cascata Photon/Nominatim (HERE_API_KEY não configurada)"
        print(f"🌍 Geocodificando {len(novos)} endereço(s) novo(s)/desatualizado(s) via {motor}...")
        if not _HERE_API_KEY:
            print("   ⚠️  Defina HERE_API_KEY pra usar a camada mais precisa (opcional).")

    for i, (cep, numero) in enumerate(novos, start=1):
        coord = _geocodificar_par(cep, numero, cidade, estado)
        chave = _chave_cache(cep, numero)
        rotulo_num = numero or "s/nº"
        if coord:
            lat, lon, fonte = coord
            cache_bruto[chave] = {"lat": lat, "lon": lon, "fonte": fonte}
            print(f"  [{i}/{len(novos)}] {cep} ({rotulo_num}) -> ({lat}, {lon}, {fonte}) -> {_link_maps(lat, lon)}")
        else:
            cache_bruto[chave] = None
            print(f"  [{i}/{len(novos)}] {cep} ({rotulo_num}) -> None")

    _salvar_cache(cache_bruto)

    for par in pares:
        resultado[par] = _cache_para_coord(cache_bruto.get(_chave_cache(*par)))
    return resultado


def csv_para_rota_xlsx(caminho_csv: str, caminho_xlsx: str | None = None, geocodificar: bool = True) -> str:
    caminho_csv = Path(caminho_csv)
    if not caminho_csv.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {caminho_csv}")

    with open(caminho_csv, encoding="utf-8-sig", newline="") as f:
        leitor = csv.DictReader(f)
        linhas = list(leitor)

    if not linhas:
        raise ValueError("CSV vazio.")

    # extrai endereço detalhado + CEP + número de cada linha
    processadas = []
    pares_unicos = set()
    for linha in linhas:
        address_bruto = linha.get("Address", "")
        endereco = _extrair_endereco_combinado(address_bruto)
        cep = _extrair_cep(address_bruto)
        numero = _extrair_numero(address_bruto)
        cidade = linha.get("City", "").strip()
        estado = linha.get("State", "").strip()
        contato = linha.get("Contact", "").strip()
        processadas.append((endereco, cidade, estado, contato, cep, numero))
        if cep:
            pares_unicos.add((cep, numero))

    coords_por_par = {}
    if geocodificar and pares_unicos:
        # assume cidade/estado predominantes do arquivo (normalmente é tudo a mesma cidade)
        cidade_ref = processadas[0][1] or "Goiânia"
        estado_ref = processadas[0][2] or "GO"
        coords_por_par = _geocodificar_enderecos(pares_unicos, cidade_ref, estado_ref)

    wb = Workbook()
    ws = wb.active
    ws.title = "Rota"

    # IMPORTANTE: o script de tratamento de dados (tratamento_dados.py) lê a
    # 5ª coluna (índice 4, começando do zero) especificamente para preencher
    # o campo "ENDERECO_ORIGINAL" (o texto secundário exibido embaixo do
    # endereço principal no app). Por isso repetimos o endereço detalhado
    # nessa posição — senão esse campo ficaria com o telefone (Contact) no
    # lugar do endereço.
    cabecalho = ["Sequence", "Address", "City", "State", "Address (Original)", "Contact", "Lat", "Lon", "Geo Fonte", "Link Mapa"]
    fundo = PatternFill("solid", fgColor="2E4057")
    fonte_cabecalho = Font(color="FFFFFF", bold=True)
    for col, titulo in enumerate(cabecalho, start=1):
        cel = ws.cell(row=1, column=col, value=titulo)
        cel.fill = fundo
        cel.font = fonte_cabecalho

    fonte_link = Font(color="1155CC", underline="single")

    contagem_fontes: dict = {}

    for i, (endereco, cidade, estado, contato, cep, numero) in enumerate(processadas, start=1):
        lat, lon, fonte_geo = (None, None, None)
        if cep:
            par_coord = coords_por_par.get((cep, numero))
            if par_coord:
                lat, lon, fonte_geo = par_coord
                contagem_fontes[fonte_geo] = contagem_fontes.get(fonte_geo, 0) + 1

        rotulo_fonte = _FONTE_LABELS.get(fonte_geo, "Sem coordenada")

        ws.cell(row=i + 1, column=1, value=i)
        ws.cell(row=i + 1, column=2, value=endereco)
        ws.cell(row=i + 1, column=3, value=cidade)
        ws.cell(row=i + 1, column=4, value=estado)
        ws.cell(row=i + 1, column=5, value=endereco)
        ws.cell(row=i + 1, column=6, value=contato)
        ws.cell(row=i + 1, column=7, value=lat)
        ws.cell(row=i + 1, column=8, value=lon)
        ws.cell(row=i + 1, column=9, value=rotulo_fonte)

        if lat is not None and lon is not None:
            cel_link = ws.cell(row=i + 1, column=10, value="Abrir no Maps")
            cel_link.hyperlink = _link_maps(lat, lon)
            cel_link.font = fonte_link

    larguras = [10, 55, 16, 8, 55, 18, 12, 12, 26, 16]
    for col, largura in enumerate(larguras, start=1):
        ws.column_dimensions[chr(64 + col)].width = largura

    if caminho_xlsx is None:
        caminho_xlsx = caminho_csv.with_suffix(".xlsx")
    else:
        caminho_xlsx = Path(caminho_xlsx)

    wb.save(caminho_xlsx)

    sem_cep = sum(1 for _, _, _, _, cep, _ in processadas if not cep)
    sem_coord = sum(
        1 for _, _, _, _, cep, numero in processadas
        if cep and not coords_por_par.get((cep, numero))
    )
    print(f"✅ XLSX gerado em: {caminho_xlsx}")
    print(f"   Linhas: {len(processadas)} | Sem CEP identificado: {sem_cep} | Sem coordenada: {sem_coord}")
    print("   Geo Fonte:")
    for fonte, rotulo in _FONTE_LABELS.items():
        qtd = contagem_fontes.get(fonte, 0)
        if qtd:
            print(f"     {rotulo}: {qtd}")

    return str(caminho_xlsx)


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python csv_para_rota_xlsx.py entrada.csv [saida.xlsx] [--no-geo]")
        sys.exit(1)

    args = [a for a in sys.argv[1:] if a != "--no-geo"]
    geocodificar = "--no-geo" not in sys.argv

    entrada = args[0]
    saida = args[1] if len(args) > 1 else None

    csv_para_rota_xlsx(entrada, saida, geocodificar=geocodificar)
