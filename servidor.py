"""
servidor.py  —  Rota Manager  (FastAPI)
========================================
Substitui o http.server original por FastAPI + Uvicorn.
Todas as rotas, lógica de negócio, planos e integrações
foram preservadas 100%.  O frontend (rota_manager1.html)
não precisa de nenhuma alteração.

Rodando localmente:
    pip install fastapi uvicorn python-multipart requests openpyxl
    python servidor.py

Railway / Render / Fly:
    A plataforma define PORT via variável de ambiente.
    Configure também: HERE_API_KEY, BREVO_API_KEY, BREVO_SENDER_EMAIL,
    DATA_DIR, ADMIN_PASS, ORS_API_KEY, GOOGLE_GEOCODING_API_KEY
    (opcionais mas recomendados em produção).

    ORS_API_KEY: chave gratuita do OpenRouteService (openrouteservice.org),
    usada em /rota/otimizar. Sem ela, cai automaticamente para o servidor
    público de demonstração do OSRM, que não tem SLA e pode falhar.
"""

import hashlib
import json
import os
import re
import secrets
import subprocess
import sys
import threading
import uuid
from datetime import datetime, timedelta
from pathlib import Path

import requests
import uvicorn
from fastapi import FastAPI, Request, Header, HTTPException, UploadFile, File
from fastapi.responses import JSONResponse, FileResponse, Response
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles

import gamification

# ═══════════════════════════════════════════════════════════════════
#  CONFIGURAÇÃO
# ═══════════════════════════════════════════════════════════════════

HOST           = os.environ.get("HOST", "0.0.0.0")
PORT           = int(os.environ.get("PORT", 8792))
HTML_FILE      = "rota_manager1.html"
ARQ_ENTRADA    = "rota.xlsx"
ARQ_PROCESSADO = "rota_processada_final.xlsx"
ARQ_VALIDADO   = "rota_validada_here.xlsx"
TRATAMENTO_PY  = "tratamento_dados.py"
ANJUN_SCRIPT   = "csv_para_rota_xlsx.py"

HERE_API_KEY   = os.environ.get("HERE_API_KEY", "P8C0izk0pJ1PIZr3d5CpeAI8b_dc7YFLkNKJlzP0A-M").strip()
HERE_CIDADE_UF = os.environ.get("HERE_CIDADE_UF", "Goiânia - GO, Brasil").strip()

OSRM_BASE_URL  = os.environ.get("OSRM_BASE_URL", "https://router.project-osrm.org")

ORS_API_KEY    = os.environ.get("ORS_API_KEY", "").strip()
ORS_BASE_URL   = "https://api.openrouteservice.org"

GOOGLE_VISION_API_KEY = (os.environ.get("GOOGLE_VISION_API_KEY") or "").strip() or None
GOOGLE_VISION_URL     = "https://vision.googleapis.com/v1/images:annotate"

# Usada só como fallback de geocodificação (depois que o HERE falha), e só
# pra endereços "simples" (rua + número), pra economizar a cota. Configure
# GOOGLE_GEOCODING_API_KEY no Railway pra ativar; sem ela, o app segue só
# com HERE + banco de coordenadas normalmente.
GOOGLE_GEOCODING_API_KEY = os.environ.get("GOOGLE_GEOCODING_API_KEY", "").strip()
GOOGLE_GEOCODING_URL     = "https://maps.googleapis.com/maps/api/geocode/json"

BREVO_API_URL       = "https://api.brevo.com/v3/smtp/email"
BREVO_API_KEY       = os.environ.get("BREVO_API_KEY")
BREVO_SENDER_EMAIL  = os.environ.get("BREVO_SENDER_EMAIL")
BREVO_SENDER_NOME   = os.environ.get("BREVO_SENDER_NOME", "Rota Manager")
EMAIL_TTL_MINUTOS   = 10

DATA_DIR = Path(os.environ.get("DATA_DIR", "/data" if Path("/data").is_dir() else "."))
DATA_DIR.mkdir(parents=True, exist_ok=True)

USERS_FILE        = str(DATA_DIR / "usuarios.json")
HISTORICO_FILE    = str(DATA_DIR / "historico_rotas.json")
BANCO_COORDS_FILE = str(DATA_DIR / "banco_coords.json")

PAGAMENTO_LINK_REUSE_MINUTOS  = 30
INFINITEPAY_HANDLE            = "moisessenju"
INFINITEPAY_LINKS_URL         = "https://api.checkout.infinitepay.io/links"
INFINITEPAY_PAYMENT_CHECK_URL = "https://api.checkout.infinitepay.io/payment_check"

# ─── Planos ────────────────────────────────────────────────────────
PLANOS = {
    "avulsa": {
        "nome":      "Importação Avulsa",
        "preco":     2.00,
        "tipo":      "avulso",
        "dias":      None,
        "beneficio": "Use 1 importação avulsa",
        "badge":     None,
        "pagamento_automatico": True,
        "importacoes_por_dia": None,
    },
    "essencial": {
        "nome":      "Plano Essencial",
        "preco":     30.00,
        "tipo":      "mensal",
        "dias":      30,
        "beneficio": "1 importação por dia",
        "badge":     None,
        "pagamento_automatico": True,
        "importacoes_por_dia": 1,
    },
    "profissional": {
        "nome":      "Plano Profissional",
        "preco":     60.00,
        "tipo":      "mensal",
        "dias":      30,
        "beneficio": "2 importações por dia",
        "badge":     "MAIS POPULAR",
        "pagamento_automatico": True,
        "importacoes_por_dia": 2,
    },
}

SESSION_TTL_HORAS = 12

# ═══════════════════════════════════════════════════════════════════
#  FASTAPI APP
# ═══════════════════════════════════════════════════════════════════

app = FastAPI(title="Rota Manager", docs_url=None, redoc_url=None)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["GET", "POST", "DELETE", "OPTIONS"],
    allow_headers=["*"],
)

# ─── Gamificação ────────────────────────────────────────────────────
# (o include_router() fica lá embaixo, DEPOIS da rota /api/perfil/me —
# senão a rota genérica /api/perfil/{user_id} do gamification.router
# "rouba" a palavra "me" como se fosse um user_id)

_ICONS_DIR = Path(__file__).parent / "static" / "icons"
if _ICONS_DIR.is_dir():
    app.mount("/static/icons", StaticFiles(directory=str(_ICONS_DIR)), name="icons")


def _garantir_perfil(user_id: str) -> "gamification.PerfilUsuario":
    """Cria o perfil de gamificação na primeira vez que o usuário aparece."""
    perfil = gamification.carregar_usuario(user_id)
    if perfil is None:
        perfil = gamification.PerfilUsuario(user_id=user_id)
        gamification.salvar_usuario(perfil)
    return perfil


def _conceder_xp_rota(user_id: str, paradas: int, pacotes: int, rota_hash: str) -> dict | None:
    """Concede XP por rota concluída. Nunca deixa um erro de gamificação
    derrubar o pipeline principal — só loga e segue."""
    try:
        _garantir_perfil(user_id)
        payload = gamification.RegistrarRotaPayload(
            user_id=user_id, paradas_concluidas=paradas,
            pacotes_entregues=pacotes, rota_hash=rota_hash,
        )
        resultado = gamification.registrar_conclusao_rota(payload)
        if resultado.xp_ganho:
            print(f"  [XP] {user_id} +{resultado.xp_ganho} XP (rota) "
                  f"— nível {resultado.perfil.nivel}"
                  f"{' 🎉 subiu de nível!' if resultado.subiu_de_nivel else ''}")
        return resultado.model_dump()
    except Exception as e:
        print(f"  [XP] ⚠️ falha ao conceder XP de rota pra {user_id}: {e}")
        return None


def _conceder_xp_endereco(user_id: str, endereco: str) -> dict | None:
    try:
        _garantir_perfil(user_id)
        payload = gamification.RegistrarEnderecoPayload(user_id=user_id, enderecos=[endereco])
        resultado = gamification.registrar_endereco_corrigido(payload)
        if resultado.xp_ganho:
            print(f"  [XP] {user_id} +{resultado.xp_ganho} XP (endereço) "
                  f"— nível {resultado.perfil.nivel}"
                  f"{' 🎉 subiu de nível!' if resultado.subiu_de_nivel else ''}")
        return resultado.model_dump()
    except Exception as e:
        print(f"  [XP] ⚠️ falha ao conceder XP de endereço pra {user_id}: {e}")
        return None

def ok_json(data: dict, status: int = 200) -> JSONResponse:
    return JSONResponse(content=data, status_code=status)

def err_json(msg: str, status: int = 400) -> JSONResponse:
    return JSONResponse(content={"ok": False, "erro": msg}, status_code=status)

def _base_url(request: Request) -> str:
    host = request.headers.get("x-forwarded-host") or request.headers.get("host") or f"localhost:{PORT}"
    if host.split(":")[0] in ("localhost", "127.0.0.1"):
        proto = "http"
    else:
        proto = request.headers.get("x-forwarded-proto", "https")
    return f"{proto}://{host}"

# ═══════════════════════════════════════════════════════════════════
#  MIGRAÇÃO DE VOLUME
# ═══════════════════════════════════════════════════════════════════

def _migrar_para_volume():
    if str(DATA_DIR) == ".":
        return
    for nome in ("usuarios.json", "historico_rotas.json", "banco_coords.json"):
        destino = DATA_DIR / nome
        origem  = Path(nome)
        if not destino.exists() and origem.exists():
            try:
                destino.write_text(origem.read_text("utf-8"), "utf-8")
                print(f"  [migração] {nome} → volume ({DATA_DIR})")
            except Exception as e:
                print(f"  [migração] falha {nome}: {e}")

# ═══════════════════════════════════════════════════════════════════
#  SESSÕES EM MEMÓRIA
# ═══════════════════════════════════════════════════════════════════

_sessoes: dict = {}

def _limpar_sessoes_expiradas():
    agora = datetime.now()
    expiradas = [t for t, s in _sessoes.items()
                 if agora - s["criado_em"] > timedelta(hours=SESSION_TTL_HORAS)]
    for t in expiradas:
        del _sessoes[t]

def criar_sessao(user_id: str, usuario: str, is_admin: bool = False) -> str:
    _limpar_sessoes_expiradas()
    token = secrets.token_hex(32)
    _sessoes[token] = {
        "user_id":   user_id,
        "usuario":   usuario,
        "is_admin":  is_admin,
        "dados":     None,
        "criado_em": datetime.now(),
    }
    return token

def obter_sessao(token: str) -> dict | None:
    _limpar_sessoes_expiradas()
    s = _sessoes.get(token)
    if s is None:
        return None
    if datetime.now() - s["criado_em"] > timedelta(hours=SESSION_TTL_HORAS):
        del _sessoes[token]
        return None
    return s

def destruir_sessao(token: str):
    _sessoes.pop(token, None)

def _token_da_request(request: Request) -> str | None:
    auth = request.headers.get("authorization", "")
    if auth.startswith("Bearer "):
        return auth[7:].strip()
    return None

def _sessao_ou_401(request: Request) -> dict:
    token = _token_da_request(request)
    if not token:
        raise HTTPException(status_code=401, detail="Não autenticado.")
    sess = obter_sessao(token)
    if sess is None:
        raise HTTPException(status_code=401, detail="Sessão expirada ou inválida. Faça login novamente.")
    return sess

def _sessao_admin_ou_403(request: Request) -> dict:
    sess = _sessao_ou_401(request)
    if not sess.get("is_admin"):
        raise HTTPException(status_code=403, detail="Acesso restrito ao administrador.")
    return sess

def _sessao_com_acesso_ou_403(request: Request) -> dict:
    sess = _sessao_ou_401(request)
    if sess.get("is_admin"):
        return sess
    if not usuario_tem_acesso_ativo(sess["usuario"]):
        raise HTTPException(status_code=403,
            detail="Seu acesso à importação de rotas expirou ou não foi liberado. Fale com o administrador.")
    pode, motivo = usuario_pode_importar_hoje(sess["usuario"])
    if not pode:
        raise HTTPException(status_code=403, detail=motivo)
    return sess

# ─── Converte HTTPException em JSON padrão do app ──────────────────
@app.exception_handler(HTTPException)
async def http_exception_handler(request: Request, exc: HTTPException):
    return JSONResponse(
        status_code=exc.status_code,
        content={"ok": False, "erro": exc.detail},
    )

# ═══════════════════════════════════════════════════════════════════
#  EMAIL  (Brevo)
# ═══════════════════════════════════════════════════════════════════

def _email_valido(email: str) -> bool:
    return bool(re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", email or ""))

def _gerar_codigo() -> str:
    return f"{secrets.randbelow(1_000_000):06d}"

def enviar_codigo_email(destino: str, codigo: str) -> tuple[bool, str]:
    if not BREVO_API_KEY or not BREVO_SENDER_EMAIL:
        return False, "Servidor não configurado para enviar email (BREVO_API_KEY/BREVO_SENDER_EMAIL ausentes)."
    try:
        payload = {
            "sender":      {"name": BREVO_SENDER_NOME, "email": BREVO_SENDER_EMAIL},
            "to":          [{"email": destino}],
            "subject":     "Seu código de verificação — Rota Manager",
            "textContent": (
                f"Seu código de verificação do Rota Manager é: {codigo}\n\n"
                f"Esse código expira em {EMAIL_TTL_MINUTOS} minutos.\n"
                f"Se você não solicitou este cadastro, ignore este email."
            ),
        }
        headers = {"api-key": BREVO_API_KEY, "Content-Type": "application/json", "Accept": "application/json"}
        r = requests.post(BREVO_API_URL, json=payload, headers=headers, timeout=15)
        if r.status_code in (200, 201):
            return True, ""
        return False, f"Brevo retornou erro {r.status_code}: {r.text[:200]}"
    except Exception as e:
        return False, f"Falha ao enviar email: {e}"

# ═══════════════════════════════════════════════════════════════════
#  CADASTRO PENDENTE (verificação por email)
# ═══════════════════════════════════════════════════════════════════

_cadastros_pendentes: dict = {}

def _limpar_cadastros_expirados():
    agora = datetime.now()
    expirados = [t for t, c in _cadastros_pendentes.items()
                 if agora - c["criado_em"] > timedelta(minutes=EMAIL_TTL_MINUTOS)]
    for t in expirados:
        del _cadastros_pendentes[t]

def _telefone_valido(tel: str) -> bool:
    digits = re.sub(r"[\s\-().]+", "", tel or "")
    return bool(re.match(r"^\d{2}9\d{8}$", digits))

def _normalizar_telefone(tel: str) -> str:
    return re.sub(r"[\s\-().]+", "", tel or "")

def iniciar_cadastro_pendente(username: str, email: str, senha: str, telefone: str = "") -> tuple[bool, str, str | None]:
    _limpar_cadastros_expirados()
    username = username.strip()
    email    = email.strip().lower()
    telefone = _normalizar_telefone(telefone) if telefone else ""

    if not username or len(username) < 3:
        return False, "Usuário deve ter pelo menos 3 caracteres.", None
    if not _email_valido(email):
        return False, "Email inválido.", None
    if not senha or len(senha) < 4:
        return False, "Senha deve ter pelo menos 4 caracteres.", None
    if telefone and not _telefone_valido(telefone):
        return False, "Telefone inválido. Use DDD + 9 + número (ex: 62 9 91153473).", None

    users = carregar_usuarios()
    chave_existente, _ = _buscar_usuario(users, username)
    if chave_existente is not None:
        return False, "Usuário já existe.", None
    if any(u.get("email", "").lower() == email for u in users.values()):
        return False, "Este email já está cadastrado em outra conta.", None

    codigo = _gerar_codigo()
    ok, erro = enviar_codigo_email(email, codigo)
    if not ok:
        return False, erro, None

    pending_token = secrets.token_hex(16)
    _cadastros_pendentes[pending_token] = {
        "username":   username,
        "email":      email,
        "telefone":   telefone,
        "senha_hash": _hash_senha(senha),
        "codigo":     codigo,
        "tentativas": 0,
        "criado_em":  datetime.now(),
    }
    return True, "Código enviado para o email.", pending_token

def confirmar_cadastro(pending_token: str, codigo: str) -> tuple[bool, str]:
    _limpar_cadastros_expirados()
    pend = _cadastros_pendentes.get(pending_token)
    if pend is None:
        return False, "Cadastro expirado ou inválido. Solicite um novo código."
    pend["tentativas"] += 1
    if pend["tentativas"] > 5:
        del _cadastros_pendentes[pending_token]
        return False, "Muitas tentativas incorretas. Solicite um novo código."
    if codigo.strip() != pend["codigo"]:
        return False, "Código incorreto."
    users = carregar_usuarios()
    chave_existente, _ = _buscar_usuario(users, pend["username"])
    if chave_existente is not None:
        del _cadastros_pendentes[pending_token]
        return False, "Usuário já existe."
    users[pend["username"]] = {
        "id":       str(uuid.uuid4()),
        "hash":     pend["senha_hash"],
        "email":    pend["email"],
        "telefone": pend.get("telefone", ""),
    }
    salvar_usuarios(users)
    del _cadastros_pendentes[pending_token]
    return True, "Conta criada com sucesso."

# ═══════════════════════════════════════════════════════════════════
#  RECUPERAÇÃO DE SENHA
# ═══════════════════════════════════════════════════════════════════

_recuperacoes_pendentes: dict = {}

def _limpar_recuperacoes_expiradas():
    agora = datetime.now()
    expirados = [t for t, c in _recuperacoes_pendentes.items()
                 if agora - c["criado_em"] > timedelta(minutes=EMAIL_TTL_MINUTOS)]
    for t in expirados:
        del _recuperacoes_pendentes[t]

def _buscar_usuario_por_login_ou_email(users: dict, identificador: str):
    alvo = (identificador or "").strip().lower()
    if not alvo:
        return None, None
    chave, dados = _buscar_usuario(users, identificador)
    if dados is not None:
        return chave, dados
    for chave, dados in users.items():
        if dados.get("email", "").lower() == alvo:
            return chave, dados
    return None, None

def iniciar_recuperacao_senha(identificador: str) -> tuple[bool, str, str | None]:
    _limpar_recuperacoes_expiradas()
    users = carregar_usuarios()
    chave, u = _buscar_usuario_por_login_ou_email(users, identificador)
    if u is None:
        return False, "Usuário ou email não encontrado.", None
    email = u.get("email", "").strip()
    if not email:
        return False, ("Esta conta não possui email cadastrado para recuperação automática. "
                       "Peça ao administrador para redefinir sua senha."), None
    codigo = _gerar_codigo()
    ok, erro = enviar_codigo_email(email, codigo)
    if not ok:
        return False, erro, None
    recovery_token = secrets.token_hex(16)
    _recuperacoes_pendentes[recovery_token] = {
        "username":   chave,
        "email":      email,
        "codigo":     codigo,
        "tentativas": 0,
        "criado_em":  datetime.now(),
    }
    em_user, _, em_dom = email.partition("@")
    mascarado = (em_user[:2] + "***@" + em_dom) if len(em_user) > 2 else ("***@" + em_dom)
    return True, mascarado, recovery_token

def confirmar_codigo_recuperacao(recovery_token: str, codigo: str) -> tuple[bool, str]:
    _limpar_recuperacoes_expiradas()
    pend = _recuperacoes_pendentes.get(recovery_token)
    if pend is None:
        return False, "Solicitação expirada ou inválida. Comece novamente."
    pend["tentativas"] += 1
    if pend["tentativas"] > 5:
        del _recuperacoes_pendentes[recovery_token]
        return False, "Muitas tentativas incorretas. Solicite um novo código."
    if codigo.strip() != pend["codigo"]:
        return False, "Código incorreto."
    pend["confirmado"] = True
    return True, "Código confirmado."

def redefinir_senha_recuperacao(recovery_token: str, nova_senha: str) -> tuple[bool, str]:
    _limpar_recuperacoes_expiradas()
    pend = _recuperacoes_pendentes.get(recovery_token)
    if pend is None:
        return False, "Solicitação expirada ou inválida. Comece novamente."
    if not pend.get("confirmado"):
        return False, "Confirme o código antes de definir a nova senha."
    if not nova_senha or len(nova_senha) < 4:
        return False, "A nova senha deve ter pelo menos 4 caracteres."
    users = carregar_usuarios()
    chave, u = _buscar_usuario(users, pend["username"])
    if u is None:
        del _recuperacoes_pendentes[recovery_token]
        return False, "Usuário não encontrado."
    u["hash"] = _hash_senha(nova_senha)
    salvar_usuarios(users)
    del _recuperacoes_pendentes[recovery_token]
    return True, "Senha redefinida com sucesso. Faça login com a nova senha."

# ═══════════════════════════════════════════════════════════════════
#  HISTÓRICO
# ═══════════════════════════════════════════════════════════════════

def carregar_historico() -> list:
    p = Path(HISTORICO_FILE)
    if not p.exists():
        return []
    try:
        return json.loads(p.read_text("utf-8"))
    except Exception:
        return []

def salvar_historico(historico: list):
    Path(HISTORICO_FILE).write_text(
        json.dumps(historico, ensure_ascii=False, indent=2), "utf-8"
    )

def adicionar_ao_historico(nome_arquivo: str, rows: list, headers: list, user_id: str = ""):
    historico = carregar_historico()
    existia_antes = any(h.get("nome") == nome_arquivo and h.get("user_id") == user_id for h in historico)
    entrada = {
        "nome":     nome_arquivo,
        "total":    len(rows),
        "headers":  headers,
        "rows":     rows,
        "user_id":  user_id,
        "salvo_em": datetime.now().strftime("%d/%m/%Y %H:%M"),
    }
    historico = [h for h in historico if not (h.get("nome") == nome_arquivo and h.get("user_id") == user_id)]
    historico.insert(0, entrada)
    historico = historico[:50]
    salvar_historico(historico)
    if existia_antes:
        # A rota anterior desse usuário saiu do histórico (foi substituída
        # por essa nova importação) — as correções manuais dele eram só
        # pra ela, então voltam ao valor global normal.
        banco_coords_limpar_overrides_usuario(user_id)
    return entrada

def atualizar_rows_historico(nome_arquivo: str, rows: list, user_id: str = "") -> bool:
    """Atualiza (in-place) as rows de uma entrada JÁ existente no histórico,
    sem disparar a limpeza de overrides — usado só pra refletir, depois do
    /pipeline, os endereços que a confirmação de geolocalização (HERE/
    Google) achou. Não conta como "a rota saiu do histórico"."""
    historico = carregar_historico()
    for h in historico:
        if h.get("nome") == nome_arquivo and h.get("user_id") == user_id:
            h["rows"]  = rows
            h["total"] = len(rows)
            salvar_historico(historico)
            return True
    return False

# ═══════════════════════════════════════════════════════════════════
#  BANCO DE COORDENADAS
# ═══════════════════════════════════════════════════════════════════

def _normalizar_endereco(end: str) -> str:
    return re.sub(r"\s+", " ", (end or "").strip().lower())

def banco_coords_carregar() -> dict:
    """
    Formato do arquivo:
        {
          "global":    { chave: {lat, lon, endereco_original, salvo_em, fonte, usuario} },
          "overrides": { user_id: { chave: {lat, lon, endereco_original, salvo_em} } }
        }
    "global" é o banco confirmado por API (HERE/Google) — permanente e
    compartilhado por todo mundo, o "ciclo" que vai pegando os prédios de
    Goiânia aos poucos. "overrides" é a correção manual de um usuário
    específico (arrastou o pin no mapa) — vale só enquanto a rota que
    contém aquele endereço estiver no histórico dele; quando ela sai do
    histórico (apagada ou substituída por nova importação), o override
    é descartado e ele volta a ver o valor global normal.
    """
    p = Path(BANCO_COORDS_FILE)
    if not p.exists():
        return {"global": {}, "overrides": {}}
    try:
        banco = json.loads(p.read_text("utf-8"))
    except Exception:
        return {"global": {}, "overrides": {}}
    # Migração do formato antigo (dict plano na raiz, sem separar override).
    if "global" not in banco and "overrides" not in banco:
        banco = {"global": banco, "overrides": {}}
    banco.setdefault("global", {})
    banco.setdefault("overrides", {})
    return banco

def banco_coords_salvar(banco: dict):
    Path(BANCO_COORDS_FILE).write_text(
        json.dumps(banco, ensure_ascii=False, indent=2), "utf-8"
    )

def banco_coords_salvar_coord(endereco: str, lat: float, lon: float, user_id: str) -> tuple[bool, str, dict]:
    """Correção manual (usuário arrastou o pin no mapa) — vale só pra ele,
    nunca sobrescreve o valor global confirmado por API. Não é permanente:
    fica só enquanto a rota que contém esse endereço estiver no histórico
    dele (veja banco_coords_limpar_overrides_usuario)."""
    chave = _normalizar_endereco(endereco)
    if not chave:
        return False, "Endereço vazio.", {}
    try:
        lat = float(lat)
        lon = float(lon)
    except (TypeError, ValueError):
        return False, "Coordenadas inválidas.", {}
    agora = datetime.now().strftime("%d/%m/%Y %H:%M")
    banco = banco_coords_carregar()
    user_id = user_id or "(desconhecido)"
    overrides_usuario = banco["overrides"].setdefault(user_id, {})
    entrada = overrides_usuario.get(chave) or {"endereco_original": endereco.strip()}
    entrada["lat"] = round(lat, 6)
    entrada["lon"] = round(lon, 6)
    entrada["endereco_original"] = entrada.get("endereco_original") or endereco.strip()
    entrada["salvo_em"] = agora
    overrides_usuario[chave] = entrada
    banco_coords_salvar(banco)
    print(f"  [BANCO_COORDS] override de {user_id!r} em {chave!r} → ({lat:.6f}, {lon:.6f})")
    return True, "Coordenada salva (só pra você, enquanto essa rota estiver no seu histórico).", {"lat": entrada["lat"], "lon": entrada["lon"]}

def banco_coords_salvar_global(endereco: str, lat: float, lon: float, fonte: str, usuario: str = "") -> dict:
    """Coordenada confirmada por API (HERE ou Google) — fica permanente no
    banco compartilhado. Só muda de novo se um usuário der override nela."""
    chave = _normalizar_endereco(endereco)
    if not chave:
        return {}
    agora = datetime.now().strftime("%d/%m/%Y %H:%M")
    banco = banco_coords_carregar()
    entrada = banco["global"].get(chave) or {"endereco_original": endereco.strip()}
    entrada["lat"] = round(float(lat), 6)
    entrada["lon"] = round(float(lon), 6)
    entrada["endereco_original"] = entrada.get("endereco_original") or endereco.strip()
    entrada["salvo_em"] = agora
    entrada["fonte"]    = fonte
    if usuario:
        entrada["usuario"] = usuario
    banco["global"][chave] = entrada
    banco_coords_salvar(banco)
    print(f"  [BANCO_COORDS] confirmado via {fonte} → {chave!r} = ({entrada['lat']:.6f}, {entrada['lon']:.6f})")
    return {"lat": entrada["lat"], "lon": entrada["lon"]}

def banco_coords_buscar(endereco: str, user_id: str = "") -> dict | None:
    """Prioriza o override pessoal do usuário (se a rota ainda estiver no
    histórico dele); senão usa o valor global confirmado por API. Retorna
    None se não achar nada — aí sim vale a pena chamar HERE/Google."""
    chave = _normalizar_endereco(endereco)
    if not chave:
        return None
    banco = banco_coords_carregar()
    if user_id:
        override = banco["overrides"].get(user_id, {}).get(chave)
        if override:
            return {"lat": override["lat"], "lon": override["lon"], "fonte": "override"}
    entrada = banco["global"].get(chave)
    if entrada:
        return {"lat": entrada["lat"], "lon": entrada["lon"], "fonte": entrada.get("fonte", "banco")}
    return None

def banco_coords_limpar_overrides_usuario(user_id: str):
    """Chamado quando a rota que tinha as correções manuais sai do
    histórico do usuário (foi apagada ou substituída por uma nova
    importação) — as correções eram só pra aquela rota, então o usuário
    volta a ver o valor global normal."""
    if not user_id:
        return
    banco = banco_coords_carregar()
    if banco["overrides"].pop(user_id, None) is not None:
        banco_coords_salvar(banco)
        print(f"  [BANCO_COORDS] overrides de {user_id!r} limpos (rota saiu do histórico)")

def banco_coords_apagar(endereco: str) -> tuple[bool, str]:
    """Remove do banco global — usado só pelo painel admin."""
    chave = _normalizar_endereco(endereco)
    banco = banco_coords_carregar()
    if chave not in banco["global"]:
        return False, "Endereço não encontrado no banco."
    del banco["global"][chave]
    banco_coords_salvar(banco)
    return True, "Entrada removida do banco."

def banco_coords_apagar_recentes(horas: float) -> int:
    """Remove do banco global toda entrada confirmada nas últimas N horas —
    usado quando uma leva de geocodificações saiu errada (ex.: HERE
    confundiu quadra/lote) e o Geni quer forçar tudo a ser reconfirmado."""
    limite = datetime.now() - timedelta(hours=horas)
    banco = banco_coords_carregar()
    removidas = []
    for chave, entrada in banco["global"].items():
        salvo_em = entrada.get("salvo_em", "")
        try:
            dt = datetime.strptime(salvo_em, "%d/%m/%Y %H:%M")
        except (ValueError, TypeError):
            continue  # sem data confiável, não mexe
        if dt >= limite:
            removidas.append(chave)
    for chave in removidas:
        del banco["global"][chave]
    if removidas:
        banco_coords_salvar(banco)
    return len(removidas)

def banco_coords_aplicar(rows: list, user_id: str = "") -> list:
    banco = banco_coords_carregar()
    overrides_usuario = banco["overrides"].get(user_id, {}) if user_id else {}
    globais = banco["global"]
    if not overrides_usuario and not globais:
        return rows
    for row in rows:
        chave = _normalizar_endereco(row.get("address", ""))
        entrada = overrides_usuario.get(chave)
        fonte = "override"
        if not entrada:
            entrada = globais.get(chave)
            fonte = entrada.get("fonte", "banco") if entrada else None
        if entrada:
            lat = str(entrada["lat"])
            lon = str(entrada["lon"])
            row["lat"]         = lat
            row["lon"]         = lon
            row["coord"]       = lat + "," + lon
            row["do_banco"]    = True
            row["fonte_coord"] = fonte
    return rows

# ─── Geocodificação sob demanda (confirmação na tela de loading) ──────

_RE_QUADRA_LOTE     = re.compile(r"\b(qd|quadra|lt|lote|q\d+|l\d+)\b", re.IGNORECASE)
_RE_NUMERO_COMPOSTO = re.compile(r"\d+\s*[-/]\s*\d+")
_RE_RUA_CODIGO      = re.compile(r"\b[A-Za-z]\s?\d{2,}\b")

def endereco_e_simples(endereco: str) -> bool:
    """
    Regra combinada com o Geni: endereço de rua com número simples
    (ex.: "RUA D, 385") x endereço com quadra/lote ou número composto
    (ex.: "RUA C152, 343-1"). Mantido pra outras decisões do fluxo; não
    decide mais nada sobre chamar o Google — isso saiu do fluxo
    automático (ver geocode_confirmar).
    """
    if not endereco:
        return False
    if _RE_QUADRA_LOTE.search(endereco):
        return False
    if _RE_NUMERO_COMPOSTO.search(endereco):
        return False
    if _RE_RUA_CODIGO.search(endereco):
        return False
    return True

# O endereço já chega padronizado do tratamento_dados.py como
# "RUA X, qd-lt", "RUA X, num" ou "RUA X, ED. Nome" (ou só "RUA X", sem
# nada). A validação estrita por resultType/houseNumber que tinha aqui
# foi removida — estava rejeitando muito endereço válido (o HERE nem
# sempre marca resultType como houseNumber mesmo quando acerta). Agora
# aceita o primeiro resultado do HERE direto, como antes.

def _here_geocode_query(query: str) -> tuple[float, float] | None:
    if not HERE_API_KEY:
        return None
    try:
        r = requests.get(
            "https://geocode.search.hereapi.com/v1/geocode",
            params={
                "q": f"{query}, {HERE_CIDADE_UF}",
                "apiKey": HERE_API_KEY,
                "limit": 1,
                "lang": "pt-BR",
            },
            timeout=8,
        )
        r.raise_for_status()
        itens = r.json().get("items") or []
        if not itens:
            return None
        pos = itens[0].get("position") or {}
        lat, lon = pos.get("lat"), pos.get("lng")
        if lat is None or lon is None:
            return None
        return round(float(lat), 6), round(float(lon), 6)
    except Exception as e:
        print(f"  [GEOCODE/HERE] falha em {query!r}: {e}")
        return None

def _google_geocode_query(query: str) -> tuple[float, float] | None:
    if not GOOGLE_GEOCODING_API_KEY:
        return None
    try:
        r = requests.get(
            GOOGLE_GEOCODING_URL,
            params={
                "address": f"{query}, {HERE_CIDADE_UF}",
                "key": GOOGLE_GEOCODING_API_KEY,
                "language": "pt-BR",
                "region": "br",
            },
            timeout=8,
        )
        r.raise_for_status()
        body = r.json()
        if body.get("status") != "OK":
            return None
        resultados = body.get("results") or []
        if not resultados:
            return None
        loc = (resultados[0].get("geometry") or {}).get("location") or {}
        lat, lon = loc.get("lat"), loc.get("lng")
        if lat is None or lon is None:
            return None
        return round(float(lat), 6), round(float(lon), 6)
    except Exception as e:
        print(f"  [GEOCODE/GOOGLE] falha em {query!r}: {e}")
        return None

# ═══════════════════════════════════════════════════════════════════
#  USUÁRIOS
# ═══════════════════════════════════════════════════════════════════

def _hash_senha(senha: str) -> str:
    return hashlib.sha256(senha.encode("utf-8")).hexdigest()

def carregar_usuarios() -> dict:
    p = Path(USERS_FILE)
    if not p.exists():
        return {}
    try:
        return json.loads(p.read_text("utf-8"))
    except Exception:
        return {}

def salvar_usuarios(users: dict):
    Path(USERS_FILE).write_text(
        json.dumps(users, ensure_ascii=False, indent=2), "utf-8"
    )

def _buscar_usuario(users: dict, username: str):
    alvo = username.strip().lower()
    for chave, dados in users.items():
        if chave.lower() == alvo:
            return chave, dados
    return None, None

def autenticar_usuario(username: str, senha: str) -> tuple[str, str] | None:
    users = carregar_usuarios()
    chave, u = _buscar_usuario(users, username)
    if u is None or u.get("hash") != _hash_senha(senha):
        return None
    if not u.get("id"):
        u["id"] = str(uuid.uuid4())
        salvar_usuarios(users)
    return u["id"], chave

def usuario_e_admin(username: str) -> bool:
    users = carregar_usuarios()
    _, u = _buscar_usuario(users, username)
    return bool(u and u.get("is_admin"))

def usuario_tem_acesso_ativo(username: str) -> bool:
    users = carregar_usuarios()
    _, u = _buscar_usuario(users, username)
    if u is None:
        return False
    expira_raw = u.get("acesso_expira_em")
    if expira_raw:
        try:
            if datetime.now() < datetime.fromisoformat(expira_raw):
                return True
        except ValueError:
            pass
    return int(u.get("avulsa_creditos", 0) or 0) > 0

def _contagem_hoje(u: dict) -> int:
    hoje = datetime.now().strftime("%Y-%m-%d")
    c = u.get("importacoes_hoje", {})
    if not isinstance(c, dict) or c.get("data") != hoje:
        return 0
    return int(c.get("count", 0) or 0)

def usuario_pode_importar_hoje(username: str) -> tuple[bool, str]:
    users = carregar_usuarios()
    _, u = _buscar_usuario(users, username)
    if u is None:
        return False, "Usuário não encontrado."
    plano_ativo = u.get("plano_ativo")
    if not plano_ativo:
        return True, ""
    limite = PLANOS.get(plano_ativo, {}).get("importacoes_por_dia")
    if limite is None:
        return True, ""
    usadas = _contagem_hoje(u)
    if usadas >= limite:
        sufixo = "ão" if limite == 1 else "ões"
        return False, (f"Limite diário do seu plano atingido "
                       f"({usadas}/{limite} importaç{sufixo} hoje). Volte amanhã.")
    return True, ""

def registrar_importacao_hoje(username: str):
    users = carregar_usuarios()
    chave, u = _buscar_usuario(users, username)
    if u is None or not u.get("plano_ativo"):
        return
    hoje = datetime.now().strftime("%Y-%m-%d")
    c = u.get("importacoes_hoje", {})
    if not isinstance(c, dict) or c.get("data") != hoje:
        c = {"data": hoje, "count": 0}
    c["count"] = int(c.get("count", 0) or 0) + 1
    u["importacoes_hoje"] = c
    salvar_usuarios(users)
    limite = PLANOS.get(u["plano_ativo"], {}).get("importacoes_por_dia")
    print(f"  [LIMITE] {chave}: {c['count']} importação(ões) hoje (limite: {limite}).")

def usuario_consumir_credito_avulso_se_necessario(username: str):
    users = carregar_usuarios()
    chave, u = _buscar_usuario(users, username)
    if u is None:
        return
    tem_acesso_mensal = False
    expira_raw = u.get("acesso_expira_em")
    if expira_raw:
        try:
            tem_acesso_mensal = datetime.now() < datetime.fromisoformat(expira_raw)
        except ValueError:
            pass
    if tem_acesso_mensal:
        return
    creditos = int(u.get("avulsa_creditos", 0) or 0)
    if creditos > 0:
        u["avulsa_creditos"] = creditos - 1
        salvar_usuarios(users)
        print(f"  [ASSINATURA] Crédito avulso consumido por \"{chave}\" (restam {creditos - 1}).")

# ═══════════════════════════════════════════════════════════════════
#  ADMIN
# ═══════════════════════════════════════════════════════════════════

def admin_listar_usuarios() -> list:
    users = carregar_usuarios()
    out = []
    for nome, dados in users.items():
        out.append({
            "usuario":             nome,
            "email":               dados.get("email", ""),
            "telefone":            dados.get("telefone", ""),
            "is_admin":            bool(dados.get("is_admin", False)),
            "acesso_expira_em":    dados.get("acesso_expira_em"),
            "avulsa_creditos":     int(dados.get("avulsa_creditos", 0) or 0),
            "plano_solicitado":    dados.get("plano_solicitado"),
            "plano_solicitado_em": dados.get("plano_solicitado_em"),
        })
    out.sort(key=lambda u: u["usuario"].lower())
    return out

def admin_criar_usuario(username: str, senha: str, email: str = "", is_admin: bool = False) -> tuple[bool, str]:
    username = (username or "").strip()
    email    = (email or "").strip().lower()
    if not username or len(username) < 3:
        return False, "Usuário deve ter pelo menos 3 caracteres."
    if not senha or len(senha) < 4:
        return False, "Senha deve ter pelo menos 4 caracteres."
    if email and not _email_valido(email):
        return False, "Email inválido."
    users = carregar_usuarios()
    chave_existente, _ = _buscar_usuario(users, username)
    if chave_existente is not None:
        return False, "Usuário já existe."
    if email and any(u.get("email", "").lower() == email for u in users.values()):
        return False, "Este email já está cadastrado em outra conta."
    novo = {"id": str(uuid.uuid4()), "hash": _hash_senha(senha)}
    if email:
        novo["email"] = email
    if is_admin:
        novo["is_admin"] = True
    users[username] = novo
    salvar_usuarios(users)
    return True, "Usuário criado com sucesso."

def admin_resetar_senha(username: str, nova_senha: str) -> tuple[bool, str]:
    if not nova_senha or len(nova_senha) < 4:
        return False, "A nova senha deve ter pelo menos 4 caracteres."
    users = carregar_usuarios()
    chave, u = _buscar_usuario(users, username)
    if u is None:
        return False, "Usuário não encontrado."
    u["hash"] = _hash_senha(nova_senha)
    salvar_usuarios(users)
    return True, "Senha redefinida com sucesso."

def admin_editar_contato(username: str, email: str = "", telefone: str = "") -> tuple[bool, str]:
    users = carregar_usuarios()
    chave, u = _buscar_usuario(users, username)
    if u is None:
        return False, "Usuário não encontrado."
    email = (email or "").strip().lower()
    if email and not _email_valido(email):
        return False, "Email inválido."
    if email and any(k.lower() != chave.lower() and d.get("email", "").lower() == email
                     for k, d in users.items()):
        return False, "Este email já está cadastrado em outra conta."
    telefone_norm = _normalizar_telefone(telefone or "")
    if telefone_norm and not _telefone_valido(telefone_norm):
        return False, "Telefone inválido. Use DDD + 9 + número (ex: 62 9 91153473)."
    if email:
        u["email"] = email
    else:
        u.pop("email", None)
    if telefone_norm:
        u["telefone"] = telefone_norm
    else:
        u.pop("telefone", None)
    salvar_usuarios(users)
    return True, "Dados atualizados com sucesso."

def admin_apagar_usuario(username: str, quem_pediu: str) -> tuple[bool, str]:
    users = carregar_usuarios()
    chave, u = _buscar_usuario(users, username)
    if u is None:
        return False, "Usuário não encontrado."
    if chave.lower() == (quem_pediu or "").strip().lower():
        return False, "Você não pode apagar sua própria conta de admin enquanto está logado nela."
    del users[chave]
    salvar_usuarios(users)
    user_id = u.get("id")
    if user_id:
        for t in [t for t, s in _sessoes.items() if s.get("user_id") == user_id]:
            del _sessoes[t]
    return True, "Usuário apagado com sucesso."

def admin_liberar_acesso(username: str, dias: int) -> tuple[bool, str]:
    try:
        dias = int(dias)
    except (TypeError, ValueError):
        return False, "Número de dias inválido."
    if dias < 1:
        return False, "Informe pelo menos 1 dia de acesso."
    users = carregar_usuarios()
    chave, u = _buscar_usuario(users, username)
    if u is None:
        return False, "Usuário não encontrado."
    expira_em = datetime.now() + timedelta(days=dias)
    u["acesso_expira_em"] = expira_em.isoformat()
    u.pop("plano_ativo", None)
    salvar_usuarios(users)
    return True, f"Acesso liberado até {expira_em.strftime('%d/%m/%Y %H:%M')}."

def admin_revogar_acesso(username: str) -> tuple[bool, str]:
    users = carregar_usuarios()
    chave, u = _buscar_usuario(users, username)
    if u is None:
        return False, "Usuário não encontrado."
    u.pop("acesso_expira_em", None)
    u.pop("plano_ativo", None)
    salvar_usuarios(users)
    return True, "Acesso revogado."

# ─── Planos ────────────────────────────────────────────────────────

def _creditar_plano(u: dict, plano_id: str) -> str:
    plano = PLANOS[plano_id]
    if plano["tipo"] == "avulso":
        u["avulsa_creditos"] = int(u.get("avulsa_creditos", 0) or 0) + 1
        u.pop("plano_ativo", None)
        return "1 crédito de importação avulsa liberado."
    expira_em = datetime.now() + timedelta(days=plano["dias"])
    u["acesso_expira_em"] = expira_em.isoformat()
    u["plano_ativo"] = plano_id
    return f"{plano['nome']} liberado até {expira_em.strftime('%d/%m/%Y %H:%M')}."

def usuario_solicitar_plano(username: str, plano_id: str) -> tuple[bool, str]:
    plano = PLANOS.get(plano_id)
    if plano is None:
        return False, "Plano inválido."
    if plano.get("pagamento_automatico"):
        return False, f"{plano['nome']} usa pagamento automático — use o botão de pagamento."
    users = carregar_usuarios()
    chave, u = _buscar_usuario(users, username)
    if u is None:
        return False, "Usuário não encontrado."
    u["plano_solicitado"]    = plano_id
    u["plano_solicitado_em"] = datetime.now().isoformat()
    u.pop("pagamento_pendente", None)
    salvar_usuarios(users)
    return True, f"Solicitação de {plano['nome']} registrada. Aguarde a liberação do administrador."

def admin_confirmar_plano(username: str) -> tuple[bool, str]:
    users = carregar_usuarios()
    chave, u = _buscar_usuario(users, username)
    if u is None:
        return False, "Usuário não encontrado."
    plano_id = u.get("plano_solicitado")
    plano    = PLANOS.get(plano_id)
    if plano is None:
        return False, "Este usuário não tem solicitação de plano pendente."
    _creditar_plano(u, plano_id)
    msg = (f"1 crédito de importação avulsa liberado para \"{chave}\"."
           if plano["tipo"] == "avulso"
           else f"{plano['nome']} liberado para \"{chave}\".")
    u.pop("plano_solicitado", None)
    u.pop("plano_solicitado_em", None)
    salvar_usuarios(users)
    return True, msg

def admin_rejeitar_plano(username: str) -> tuple[bool, str]:
    users = carregar_usuarios()
    chave, u = _buscar_usuario(users, username)
    if u is None:
        return False, "Usuário não encontrado."
    u.pop("plano_solicitado", None)
    u.pop("plano_solicitado_em", None)
    salvar_usuarios(users)
    return True, "Solicitação removida."

# ═══════════════════════════════════════════════════════════════════
#  INFINITEPAY
# ═══════════════════════════════════════════════════════════════════

def _infinitepay_gerar_link(order_nsu: str, plano: dict, redirect_url: str, webhook_url: str) -> tuple[bool, str]:
    payload = {
        "handle":       INFINITEPAY_HANDLE,
        "redirect_url": redirect_url,
        "webhook_url":  webhook_url,
        "order_nsu":    order_nsu,
        "items": [{"quantity": 1, "price": int(round(plano["preco"] * 100)), "description": plano["nome"]}],
    }
    try:
        resp = requests.post(INFINITEPAY_LINKS_URL, json=payload, timeout=10)
        data = resp.json()
    except Exception as e:
        return False, f"Não foi possível conectar à InfinitePay: {e}"
    url = data.get("url")
    if not resp.ok or not url:
        erro = data.get("message") or data.get("error") or f"Erro {resp.status_code} ao gerar o link de pagamento."
        return False, erro
    return True, url

def infinitepay_consultar_pagamento(order_nsu: str, transaction_nsu: str = "", slug: str = "") -> tuple[bool, dict | str]:
    payload = {"handle": INFINITEPAY_HANDLE, "order_nsu": order_nsu,
               "transaction_nsu": transaction_nsu, "slug": slug}
    try:
        resp = requests.post(INFINITEPAY_PAYMENT_CHECK_URL, json=payload, timeout=10)
        data = resp.json()
    except Exception as e:
        return False, f"Não foi possível consultar o pagamento: {e}"
    if not resp.ok:
        return False, data.get("message") or f"Erro {resp.status_code} ao consultar pagamento."
    return True, data

def usuario_iniciar_pagamento(username: str, plano_id: str, base_url: str) -> tuple[bool, str]:
    plano = PLANOS.get(plano_id)
    if plano is None:
        return False, "Plano inválido."
    if not plano.get("pagamento_automatico"):
        return False, f"{plano['nome']} ainda usa o fluxo de solicitação manual."
    users = carregar_usuarios()
    chave, u = _buscar_usuario(users, username)
    if u is None:
        return False, "Usuário não encontrado."
    pendente = u.get("pagamento_pendente")
    if pendente and pendente.get("plano_id") == plano_id:
        try:
            criado_em = datetime.fromisoformat(pendente["criado_em"])
            if datetime.now() - criado_em < timedelta(minutes=PAGAMENTO_LINK_REUSE_MINUTOS):
                return True, pendente["url"]
        except (KeyError, ValueError):
            pass
    order_nsu    = uuid.uuid4().hex
    redirect_url = f"{base_url}/?pagamento=retorno"
    webhook_url  = f"{base_url}/webhook/infinitepay"
    ok, resultado = _infinitepay_gerar_link(order_nsu, plano, redirect_url, webhook_url)
    if not ok:
        return False, resultado
    u["pagamento_pendente"] = {
        "plano_id":  plano_id,
        "order_nsu": order_nsu,
        "url":       resultado,
        "criado_em": datetime.now().isoformat(),
    }
    salvar_usuarios(users)
    return True, resultado

def processar_pagamento_confirmado(order_nsu: str, transaction_nsu: str = "", receipt_url: str = "") -> tuple[bool, str]:
    users = carregar_usuarios()
    chave_alvo, u_alvo = None, None
    for chave, u in users.items():
        pendente = u.get("pagamento_pendente")
        if pendente and pendente.get("order_nsu") == order_nsu:
            chave_alvo, u_alvo = chave, u
            break
    if u_alvo is None:
        return False, "Pedido não encontrado."
    plano_id = u_alvo["pagamento_pendente"].get("plano_id")
    if plano_id not in PLANOS:
        return False, "Plano do pedido não existe mais."
    detalhe = _creditar_plano(u_alvo, plano_id)
    u_alvo.pop("pagamento_pendente", None)
    u_alvo["ultimo_pagamento"] = {
        "plano_id": plano_id, "order_nsu": order_nsu,
        "transaction_nsu": transaction_nsu, "receipt_url": receipt_url,
        "pago_em": datetime.now().isoformat(),
    }
    salvar_usuarios(users)
    print(f"  [INFINITEPAY] Pagamento confirmado para \"{chave_alvo}\" (order_nsu={order_nsu}). {detalhe}")
    return True, detalhe

# ═══════════════════════════════════════════════════════════════════
#  LER PLANILHA PROCESSADA
# ═══════════════════════════════════════════════════════════════════

def ler_processado(user_id: str = ""):
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("openpyxl não instalado.")
    path = Path(ARQ_VALIDADO) if Path(ARQ_VALIDADO).exists() else Path(ARQ_PROCESSADO)
    if not path.exists():
        raise FileNotFoundError(f"{ARQ_PROCESSADO} não encontrado.")
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = [str(c.value or "").strip() for c in ws[1]]

    def find_col(pats):
        for pat in pats:
            for i, h in enumerate(headers):
                if re.search(pat, h, re.IGNORECASE):
                    return i
        return None

    col_addr    = find_col([r"destination.?address", r"reformado"])
    col_stop    = find_col([r"sequence", r"stop", r"seq"])
    col_lat     = find_col([r"\blatitude\b", r"\blat\b"])
    col_lon     = find_col([r"\blongitude\b", r"\blon\b", r"\blng\b"])
    col_coord   = find_col([r"coordenadas", r"coord"])
    col_count   = find_col([r"rotas_iguais"])
    col_stops   = find_col([r"stops do grupo"])
    col_orig    = find_col([r"endere.o_original", r"original"])
    col_membros = find_col([r"membros.?json", r"membros"])
    col_validacao_here = find_col([r"validacao_here", r"validacao.here"])

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        def g(idx):
            if idx is None or idx >= len(row):
                return ""
            v = row[idx]
            return str(v).strip() if v is not None else ""

        lat   = g(col_lat)
        lon   = g(col_lon)
        coord = g(col_coord) or (f"{lat},{lon}" if lat and lon else "")
        count = int(g(col_count) or 1)
        endereco_original = g(col_orig)

        membros = []
        membros_raw = g(col_membros)
        if membros_raw:
            try:
                parsed = json.loads(membros_raw)
                if isinstance(parsed, list):
                    membros = [
                        {"stop": str(m.get("stop", "")), "original": str(m.get("original", ""))}
                        for m in parsed if isinstance(m, dict)
                    ]
            except Exception:
                membros = []

        if not membros:
            seq_fallback   = [s.strip() for s in g(col_stop).split(",") if s.strip()]
            stops_fallback = seq_fallback or [s.strip() for s in g(col_stops).replace("Stop:", "").split(",") if s.strip()]
            origs_fallback = [o.strip() for o in endereco_original.split("|") if o.strip()]
            n = max(len(stops_fallback), len(origs_fallback), 1)
            membros = [
                {"stop": stops_fallback[i] if i < len(stops_fallback) else "",
                 "original": origs_fallback[i] if i < len(origs_fallback) else endereco_original}
                for i in range(n)
            ]

        entry = {
            "address":            g(col_addr),
            "stop":               g(col_stop),
            "lat":                lat,
            "lon":                lon,
            "coord":              coord,
            "group_size":         count,
            "group_stops":        g(col_stops),
            "group_label":        g(col_addr),
            "endereco_original":  endereco_original,
            "membros":            membros,
            "validacao_here":     g(col_validacao_here),
            "_cid":               str(uuid.uuid4()),
        }
        rows.append(entry)

    rows = banco_coords_aplicar(rows, user_id)
    return rows, headers

# ═══════════════════════════════════════════════════════════════════
#  OTIMIZAÇÃO DE ROTA (OSRM /trip/)
# ═══════════════════════════════════════════════════════════════════

def _osrm_otimizar_sequencia(coords: list[tuple[float, float]]) -> list[int] | None:
    """
    Recebe uma lista de (lat, lon) e devolve a ordem otimizada dos ÍNDICES
    originais (mesmo tamanho da entrada), usando o serviço público OSRM
    /trip/ (baseado em OpenStreetMap). O primeiro ponto da lista é fixado
    como início da rota (source=first) e não faz retorno ao ponto de
    partida (roundtrip=false).

    Retorna None se o serviço falhar ou a resposta vier inconsistente —
    nesse caso o chamador deve manter a ordem original.
    """
    if len(coords) < 2:
        return list(range(len(coords)))

    coord_str = ";".join(f"{lon:.6f},{lat:.6f}" for lat, lon in coords)
    url = f"{OSRM_BASE_URL}/trip/v1/driving/{coord_str}"
    params = {"source": "first", "roundtrip": "false", "overview": "false"}
    # O servidor de demonstração do OSRM fica atrás de uma proteção anti-bot
    # que costuma barrar requisições com o User-Agent padrão do `requests`
    # (ex: "python-requests/2.x"). Um UA de navegador evita esse bloqueio.
    headers = {
        "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                        "AppleWebKit/537.36 (KHTML, like Gecko) "
                        "Chrome/124.0.0.0 Safari/537.36"),
        "Accept": "application/json",
    }

    try:
        r = requests.get(url, params=params, headers=headers, timeout=30)
        status = r.status_code
        if status != 200:
            print(f"  [OSRM] HTTP {status} — corpo: {r.text[:300]!r}")
            return None
        data = r.json()
    except requests.exceptions.RequestException as e:
        print(f"  [OSRM] falha na chamada: {type(e).__name__}: {e}")
        return None

    if data.get("code") != "Ok":
        print(f"  [OSRM] resposta inesperada: {data.get('code')} — {data.get('message', '')}")
        return None

    waypoints = data.get("waypoints", [])
    if len(waypoints) != len(coords):
        print("  [OSRM] número de waypoints não bate com a entrada.")
        return None

    # waypoints[i]['waypoint_index'] = posição do ponto i na rota otimizada
    try:
        ordem = sorted(range(len(coords)), key=lambda i: waypoints[i]["waypoint_index"])
    except (KeyError, TypeError):
        return None
    return ordem

def _ors_otimizar_sequencia(coords: list[tuple[float, float]]) -> list[int] | None:
    """
    Mesmo contrato de _osrm_otimizar_sequencia: recebe (lat, lon) e devolve a
    ordem otimizada dos ÍNDICES originais. Usa o endpoint /optimization do
    OpenRouteService (motor VROOM), que é hospedado com SLA — muito mais
    confiável que o servidor público de demonstração do OSRM.

    Fixa o primeiro ponto da lista como início do "veículo" (não é tratado
    como job) e todos os demais pontos entram como jobs a visitar na ordem
    mais eficiente. Não exige retorno ao ponto de partida (sem "end").
    """
    if len(coords) < 2:
        return list(range(len(coords)))
    if not ORS_API_KEY:
        print("  [ORS] ORS_API_KEY não configurada — pulei para o próximo serviço.")
        return None

    inicio = coords[0]
    jobs = [
        {"id": i, "location": [lon, lat]}
        for i, (lat, lon) in enumerate(coords) if i != 0
    ]
    payload = {
        "jobs": jobs,
        "vehicles": [{
            "id": 1,
            "profile": "driving-car",
            "start": [inicio[1], inicio[0]],
        }],
    }
    headers = {
        "Authorization": ORS_API_KEY,
        "Content-Type": "application/json",
        "Accept": "application/json",
    }

    try:
        r = requests.post(f"{ORS_BASE_URL}/optimization", json=payload, headers=headers, timeout=30)
        if r.status_code != 200:
            print(f"  [ORS] HTTP {r.status_code} — corpo: {r.text[:300]!r}")
            return None
        data = r.json()
    except requests.exceptions.RequestException as e:
        print(f"  [ORS] falha na chamada: {type(e).__name__}: {e}")
        return None

    routes = data.get("routes") or []
    if not routes:
        print(f"  [ORS] resposta sem rotas: {str(data)[:300]}")
        return None

    steps = routes[0].get("steps", [])
    ordem_jobs = [s["id"] for s in steps if s.get("type") == "job"]
    if len(ordem_jobs) != len(jobs):
        print(f"  [ORS] {len(ordem_jobs)} jobs retornados, esperava {len(jobs)}.")
        return None

    return [0] + ordem_jobs

def _otimizar_sequencia(coords: list[tuple[float, float]]) -> tuple[list[int] | None, str]:
    """
    Escolhe o serviço de otimização: OpenRouteService primeiro (se a chave
    estiver configurada — é o caminho recomendado e mais estável), caindo
    para o OSRM público como último recurso caso a chave não esteja
    configurada ou o ORS falhe. Retorna (ordem, nome_do_servico_usado).
    """
    if ORS_API_KEY:
        ordem = _ors_otimizar_sequencia(coords)
        if ordem is not None:
            return ordem, "openrouteservice"
        print("  [OTIMIZAR] ORS falhou, tentando OSRM público como fallback...")

    ordem = _osrm_otimizar_sequencia(coords)
    return ordem, "osrm"

# ═══════════════════════════════════════════════════════════════════
#  BOOTSTRAP DO ADMIN
# ═══════════════════════════════════════════════════════════════════

def _bootstrap_admin():
    users = carregar_usuarios()
    chave, u = _buscar_usuario(users, "admin")
    if u is None:
        senha_inicial = os.environ.get("ADMIN_PASS", "admin123")
        users["admin"] = {
            "id":       str(uuid.uuid4()),
            "hash":     _hash_senha(senha_inicial),
            "is_admin": True,
        }
        salvar_usuarios(users)
        print(f"  [ADMIN] Usuário 'admin' criado. Senha inicial: {senha_inicial!r}")
    elif not u.get("is_admin"):
        u["is_admin"] = True
        salvar_usuarios(users)
        print("  [ADMIN] Usuário 'admin' recebeu a flag is_admin.")

# ═══════════════════════════════════════════════════════════════════
#  ROTAS  ──  GET
# ═══════════════════════════════════════════════════════════════════

@app.get("/ping")
async def ping():
    return {"ok": True}

@app.get("/")
@app.get("/index")
@app.get(f"/{HTML_FILE}")
async def serve_html():
    html_path = Path(HTML_FILE)
    if not html_path.exists():
        raise HTTPException(status_code=404, detail="rota_manager1.html não encontrado.")
    return FileResponse(html_path, media_type="text/html; charset=utf-8")

@app.get("/dados")
async def get_dados(request: Request):
    sess = _sessao_ou_401(request)
    if sess["dados"] is None:
        return err_json("Nenhum dado processado ainda.", 404)
    rows, headers = sess["dados"]
    return ok_json({"ok": True, "arquivo": ARQ_PROCESSADO, "rows": rows, "headers": headers})

@app.get("/auth/status")
async def auth_status(request: Request):
    sess = _sessao_ou_401(request)
    tem_acesso = bool(sess.get("is_admin")) or usuario_tem_acesso_ativo(sess["usuario"])
    return ok_json({"ok": True, "tem_acesso": tem_acesso, "is_admin": bool(sess.get("is_admin"))})

@app.get("/api/perfil/me")
async def perfil_gamificacao_me(request: Request):
    """Atalho pro perfil de gamificação do usuário logado (sem precisar
    que o frontend conheça o user_id — só o token de sessão)."""
    sess = _sessao_ou_401(request)
    perfil = _garantir_perfil(sess["user_id"])
    return ok_json({
        "ok": True,
        "perfil": perfil.model_dump(),
        "nome_exibicao": perfil.nome_personagem or sess["usuario"],
        "xp_para_proximo_nivel": gamification.xp_necessario_para_nivel(perfil.nivel),
        "icon_base_path": gamification.ICON_BASE_PATH,
        "badges_desbloqueadas": gamification.calcular_badges_desbloqueadas(perfil.nivel),
        "itens_desbloqueados": gamification.calcular_itens_desbloqueados(perfil),
    })

@app.post("/api/perfil/me/nome")
async def perfil_gamificacao_atualizar_nome(request: Request):
    """Define/limpa o apelido do personagem do usuário logado. Se vazio,
    o personagem volta a usar o nome de usuário normal."""
    sess = _sessao_ou_401(request)
    data = await request.json()
    nome_novo = (data.get("nome_personagem") or "").strip()
    if nome_novo:
        if len(nome_novo) < 2 or len(nome_novo) > 20:
            return err_json("O nome do personagem deve ter entre 2 e 20 caracteres.")
    perfil = _garantir_perfil(sess["user_id"])
    perfil.nome_personagem = nome_novo or None
    gamification.salvar_usuario(perfil)
    return ok_json({
        "ok": True,
        "msg": "Nome do personagem atualizado.",
        "nome_exibicao": perfil.nome_personagem or sess["usuario"],
    })

@app.get("/api/perfil/ranking")
async def perfil_ranking(request: Request):
    """Ranking de todos os jogadores por nível (desempate por XP total)."""
    sess = _sessao_ou_401(request)
    users = carregar_usuarios()
    id_para_usuario = {u.get("id"): chave for chave, u in users.items() if u.get("id")}

    ranking = []
    for p in gamification.listar_todos_perfis():
        username = id_para_usuario.get(p.user_id, "Piloto")
        ranking.append({
            "user_id":       p.user_id,
            "nome_exibicao": p.nome_personagem or username,
            "nivel":         p.nivel,
            "xp_total":      p.xp_total,
        })
    ranking.sort(key=lambda r: (-r["nivel"], -r["xp_total"]))
    for i, r in enumerate(ranking, start=1):
        r["posicao"] = i

    minha_posicao = next((r for r in ranking if r["user_id"] == sess["user_id"]), None)
    return ok_json({
        "ok": True,
        "ranking": ranking[:20],
        "minha_posicao": minha_posicao,
        "total_jogadores": len(ranking),
    })

@app.get("/auth/perfil")
async def auth_perfil(request: Request):
    sess = _sessao_ou_401(request)
    users = carregar_usuarios()
    chave, u = _buscar_usuario(users, sess["usuario"])
    if u is None:
        return err_json("Usuário não encontrado.")
    return ok_json({"ok": True, "usuario": chave, "email": u.get("email", ""), "telefone": u.get("telefone", "")})

@app.get("/planos")
async def get_planos(request: Request):
    _sessao_ou_401(request)
    planos = [{"id": pid, **dados} for pid, dados in PLANOS.items()]
    return ok_json({"ok": True, "planos": planos})

@app.get("/assinatura/status")
async def assinatura_status(request: Request):
    sess = _sessao_ou_401(request)
    users = carregar_usuarios()
    _, u = _buscar_usuario(users, sess["usuario"])
    if u is None:
        return err_json("Usuário não encontrado.", 404)
    plano_solicitado = u.get("plano_solicitado")
    pendente = u.get("pagamento_pendente")
    return ok_json({
        "ok":                    True,
        "acesso_expira_em":      u.get("acesso_expira_em"),
        "avulsa_creditos":       int(u.get("avulsa_creditos", 0) or 0),
        "plano_solicitado":      plano_solicitado,
        "plano_solicitado_em":   u.get("plano_solicitado_em"),
        "plano_solicitado_nome": PLANOS.get(plano_solicitado, {}).get("nome"),
        "plano_ativo":           u.get("plano_ativo"),
        "usadas_hoje":           _contagem_hoje(u),
        "limite_hoje":           PLANOS.get(u.get("plano_ativo", ""), {}).get("importacoes_por_dia"),
        "pagamento_pendente": {
            "plano_id":  pendente.get("plano_id"),
            "url":       pendente.get("url"),
            "criado_em": pendente.get("criado_em"),
        } if pendente else None,
    })

@app.get("/assinatura/confirmar-pagamento")
async def assinatura_confirmar_pagamento(request: Request,
                                          order_nsu: str = "",
                                          transaction_nsu: str = "",
                                          slug: str = ""):
    sess = _sessao_ou_401(request)
    if not order_nsu:
        return err_json("order_nsu ausente.", 400)
    users = carregar_usuarios()
    _, u = _buscar_usuario(users, sess["usuario"])
    pendente = u.get("pagamento_pendente") if u else None
    if not pendente or pendente.get("order_nsu") != order_nsu:
        return ok_json({"ok": True, "pago": True, "msg": "Pagamento já confirmado."})
    ok, info = infinitepay_consultar_pagamento(order_nsu, transaction_nsu, slug)
    if not ok:
        return err_json(info)
    if not info.get("paid"):
        return ok_json({"ok": True, "pago": False, "msg": "Pagamento ainda não confirmado."})
    ok2, msg2 = processar_pagamento_confirmado(order_nsu, transaction_nsu, info.get("receipt_url", ""))
    return ok_json({"ok": ok2, "pago": ok2, "msg": msg2})

@app.get("/historico")
async def get_historico(request: Request):
    sess = _sessao_ou_401(request)
    historico = carregar_historico()
    historico_user = [h for h in historico if h.get("user_id") == sess["user_id"]]
    resumo = [{"nome": h["nome"], "total": h["total"], "salvo_em": h.get("salvo_em", "")}
              for h in historico_user]
    return ok_json({"ok": True, "historico": resumo})

@app.get("/historico/carregar")
async def historico_carregar(request: Request, nome: str = ""):
    sess = _sessao_ou_401(request)
    historico = carregar_historico()
    entrada = next((h for h in historico if h.get("nome") == nome and h.get("user_id") == sess["user_id"]), None)
    if entrada:
        return ok_json({"ok": True, **entrada})
    return err_json("Rota não encontrada no histórico.", 404)

@app.get("/coords/listar")
async def coords_listar(request: Request):
    _sessao_admin_ou_403(request)
    banco = banco_coords_carregar()
    entradas = [{"chave": k, **v} for k, v in sorted(banco["global"].items())]
    return ok_json({"ok": True, "total": len(entradas), "entradas": entradas})

@app.get("/admin/usuarios")
async def admin_usuarios(request: Request):
    _sessao_admin_ou_403(request)
    return ok_json({"ok": True, "usuarios": admin_listar_usuarios()})

# ═══════════════════════════════════════════════════════════════════
#  ROTAS  ──  POST
# ═══════════════════════════════════════════════════════════════════

@app.post("/auth/cadastro")
async def auth_cadastro(request: Request):
    data = await request.json()
    ok, msg, pending_token = iniciar_cadastro_pendente(
        data.get("usuario", ""), data.get("email", ""), data.get("senha", ""), data.get("telefone", "")
    )
    resp = {"ok": ok, "msg": msg}
    if ok:
        resp["pending_token"] = pending_token
    return ok_json(resp)

@app.post("/auth/confirmar-cadastro")
async def auth_confirmar_cadastro(request: Request):
    data = await request.json()
    ok, msg = confirmar_cadastro(data.get("pending_token", ""), data.get("codigo", ""))
    return ok_json({"ok": ok, "msg": msg})

@app.post("/auth/recuperar")
async def auth_recuperar(request: Request):
    data = await request.json()
    ok, msg, recovery_token = iniciar_recuperacao_senha(data.get("identificador", ""))
    if ok:
        return ok_json({"ok": True, "email_mascarado": msg, "recovery_token": recovery_token})
    return ok_json({"ok": False, "erro": msg})

@app.post("/auth/recuperar-confirmar")
async def auth_recuperar_confirmar(request: Request):
    data = await request.json()
    ok, msg = confirmar_codigo_recuperacao(data.get("recovery_token", ""), data.get("codigo", ""))
    return ok_json({"ok": ok, "msg": msg})

@app.post("/auth/recuperar-nova-senha")
async def auth_recuperar_nova_senha(request: Request):
    data = await request.json()
    ok, msg = redefinir_senha_recuperacao(data.get("recovery_token", ""), data.get("nova_senha", ""))
    return ok_json({"ok": ok, "msg": msg})

@app.post("/auth/login")
async def auth_login(request: Request):
    data = await request.json()
    usuario = data.get("usuario", "").strip()
    senha   = data.get("senha", "")
    resultado = autenticar_usuario(usuario, senha)
    if resultado:
        user_id, usuario_original = resultado
        is_admin = usuario_e_admin(usuario_original)
        token = criar_sessao(user_id, usuario_original, is_admin)
        return ok_json({"ok": True, "token": token, "usuario": usuario_original, "is_admin": is_admin})
    return ok_json({"ok": False, "erro": "Usuário ou senha incorretos."})

@app.post("/auth/logout")
async def auth_logout(request: Request):
    token = _token_da_request(request)
    if token:
        destruir_sessao(token)
    return ok_json({"ok": True})

@app.post("/auth/perfil/atualizar")
async def auth_perfil_atualizar(request: Request):
    sess = _sessao_ou_401(request)
    data = await request.json()
    telefone_raw  = data.get("telefone", "").strip()
    telefone_norm = _normalizar_telefone(telefone_raw)
    if telefone_norm and not _telefone_valido(telefone_norm):
        return err_json("Telefone inválido. Use DDD + 9 + número (ex: 62 9 91153473).")
    email_novo = data.get("email", "").strip()
    if email_novo and "@" not in email_novo:
        return err_json("E-mail inválido.")
    users = carregar_usuarios()
    chave, u = _buscar_usuario(users, sess["usuario"])
    if u is None:
        return err_json("Usuário não encontrado.")
    if telefone_norm:
        users[chave]["telefone"] = telefone_norm
    if email_novo:
        users[chave]["email"] = email_novo
    salvar_usuarios(users)
    return ok_json({"ok": True, "msg": "Perfil atualizado com sucesso."})

@app.post("/assinatura/solicitar")
async def assinatura_solicitar(request: Request):
    sess = _sessao_ou_401(request)
    data = await request.json()
    ok, msg = usuario_solicitar_plano(sess["usuario"], data.get("plano", ""))
    return ok_json({"ok": ok, "msg": msg})

@app.post("/assinatura/pagar")
async def assinatura_pagar(request: Request):
    sess = _sessao_ou_401(request)
    data = await request.json()
    ok, resultado = usuario_iniciar_pagamento(sess["usuario"], data.get("plano", ""), _base_url(request))
    if ok:
        return ok_json({"ok": True, "url": resultado})
    return err_json(resultado)

@app.post("/webhook/infinitepay")
async def webhook_infinitepay(request: Request):
    """Chamado pela InfinitePay servidor-a-servidor (sem token de sessão)."""
    try:
        payload = await request.json()
    except Exception:
        return JSONResponse({"success": False, "message": "JSON inválido."}, status_code=400)
    order_nsu = payload.get("order_nsu", "")
    if not order_nsu:
        return JSONResponse({"success": False, "message": "order_nsu ausente."}, status_code=400)
    ok, msg = processar_pagamento_confirmado(
        order_nsu,
        payload.get("transaction_nsu", ""),
        payload.get("receipt_url", ""),
    )
    if ok or msg == "Pedido não encontrado.":
        return JSONResponse({"success": True, "message": None})
    return JSONResponse({"success": False, "message": msg}, status_code=400)

@app.post("/admin/usuarios/criar")
async def admin_usuarios_criar(request: Request):
    sess = _sessao_admin_ou_403(request)
    data = await request.json()
    ok, msg = admin_criar_usuario(
        data.get("usuario", ""), data.get("senha", ""),
        data.get("email", ""), bool(data.get("is_admin", False))
    )
    return ok_json({"ok": ok, "msg": msg})

@app.post("/admin/usuarios/resetar-senha")
async def admin_resetar_senha_route(request: Request):
    _sessao_admin_ou_403(request)
    data = await request.json()
    ok, msg = admin_resetar_senha(data.get("usuario", ""), data.get("nova_senha", ""))
    return ok_json({"ok": ok, "msg": msg})

@app.post("/admin/usuarios/editar")
async def admin_editar_route(request: Request):
    _sessao_admin_ou_403(request)
    data = await request.json()
    ok, msg = admin_editar_contato(data.get("usuario", ""), data.get("email", ""), data.get("telefone", ""))
    return ok_json({"ok": ok, "msg": msg})

@app.post("/admin/usuarios/liberar-acesso")
async def admin_liberar_route(request: Request):
    _sessao_admin_ou_403(request)
    data = await request.json()
    ok, msg = admin_liberar_acesso(data.get("usuario", ""), data.get("dias", 0))
    return ok_json({"ok": ok, "msg": msg})

@app.post("/admin/usuarios/revogar-acesso")
async def admin_revogar_route(request: Request):
    _sessao_admin_ou_403(request)
    data = await request.json()
    ok, msg = admin_revogar_acesso(data.get("usuario", ""))
    return ok_json({"ok": ok, "msg": msg})

@app.post("/admin/usuarios/confirmar-plano")
async def admin_confirmar_plano_route(request: Request):
    _sessao_admin_ou_403(request)
    data = await request.json()
    ok, msg = admin_confirmar_plano(data.get("usuario", ""))
    return ok_json({"ok": ok, "msg": msg})

@app.post("/admin/usuarios/rejeitar-plano")
async def admin_rejeitar_plano_route(request: Request):
    _sessao_admin_ou_403(request)
    data = await request.json()
    ok, msg = admin_rejeitar_plano(data.get("usuario", ""))
    return ok_json({"ok": ok, "msg": msg})

@app.post("/coords/salvar")
async def coords_salvar(request: Request):
    sess = _sessao_ou_401(request)
    data = await request.json()
    endereco = data.get("endereco", "")
    ok, msg, info = banco_coords_salvar_coord(
        endereco, data.get("lat", 0), data.get("lon", 0),
        sess["user_id"]
    )
    resposta = {"ok": ok, "msg": msg}
    if ok:
        resposta.update(info)
        xp_resultado = _conceder_xp_endereco(sess["user_id"], endereco)
        if xp_resultado:
            resposta["gamificacao"] = xp_resultado
    else:
        resposta["erro"] = msg
    return ok_json(resposta)

@app.post("/coords/apagar")
async def coords_apagar(request: Request):
    _sessao_admin_ou_403(request)
    data = await request.json()
    ok, msg = banco_coords_apagar(data.get("endereco", ""))
    return ok_json({"ok": ok, "msg": msg})

@app.post("/coords/apagar-recentes")
async def coords_apagar_recentes(request: Request):
    """Zera do banco global tudo que foi confirmado nas últimas N horas
    (padrão 24h) — usado quando uma leva de geocodificações saiu errada."""
    _sessao_admin_ou_403(request)
    data = await request.json()
    try:
        horas = float(data.get("horas", 24))
    except (TypeError, ValueError):
        horas = 24
    total = banco_coords_apagar_recentes(horas)
    return ok_json({"ok": True, "removidas": total})

@app.post("/geocode/confirmar")
async def geocode_confirmar(request: Request):
    """
    Confirmação de geolocalização (tela "preparando sua rota"). Pra cada
    endereço recebido, segue o ciclo: banco de coordenadas (override do
    Confirmação de geolocalização (tela "preparando sua rota"). Pra cada
    endereço recebido, olha só o banco de coordenadas (override do
    usuário, senão o global já confirmado). Sem HERE e sem Google nessa
    etapa automática — o HERE estava bagunçando endereço que a planilha
    já trouxe certo. O HERE continua disponível pra busca manual na tela
    de "abrir mapa" (isso é outro fluxo, no browser, não passa por aqui).
    Se não tem no banco, volta "encontrado: false" e o front mantém a
    geolocalização que já veio do tratamento_dados.py.
    """
    sess = _sessao_ou_401(request)
    user_id = sess["user_id"]           # chave do override pessoal (some quando a rota sai do histórico)
    data = await request.json()
    enderecos = data.get("enderecos") or []
    if not isinstance(enderecos, list):
        return err_json("Formato inválido: 'enderecos' deve ser uma lista.")
    enderecos = [e.strip() for e in enderecos if isinstance(e, str) and e.strip()]
    enderecos = enderecos[:300]  # limite de segurança por chamada

    resultados = {}
    for endereco in enderecos:
        cache = banco_coords_buscar(endereco, user_id)
        if cache:
            resultados[endereco] = {"encontrado": True, **cache}
        else:
            resultados[endereco] = {"encontrado": False}

    return ok_json({"ok": True, "resultados": resultados})

@app.post("/historico/atualizar-coords")
async def historico_atualizar_coords(request: Request):
    """Depois que o front confirma a geolocalização (HERE/Google) por cima
    dos dados que o /pipeline devolveu, manda as linhas atualizadas de
    volta aqui — pra refletir tanto no /dados desta sessão quanto no
    histórico já salvo. Sem isso, recarregar a rota (F5 ou pelo histórico)
    mostraria os badges como se nada tivesse sido confirmado."""
    sess = _sessao_ou_401(request)
    data = await request.json()
    rows = data.get("rows") or []
    if not isinstance(rows, list):
        return err_json("Formato inválido: 'rows' deve ser uma lista.")

    headers = sess["dados"][1] if sess.get("dados") else []
    sess["dados"] = (rows, headers)

    arq_final = ARQ_VALIDADO if Path(ARQ_VALIDADO).exists() else ARQ_PROCESSADO
    nome = Path(arq_final).name
    atualizado = atualizar_rows_historico(nome, rows, sess["user_id"])
    return ok_json({"ok": True, "historico_atualizado": atualizado})

@app.post("/upload")
async def upload(request: Request, arquivo: UploadFile = File(...)):
    """
    Recebe o arquivo xlsx via multipart/form-data.
    FastAPI + python-multipart fazem o parsing automaticamente —
    sem mais parsing manual de boundary.
    """
    sess = _sessao_com_acesso_ou_403(request)
    contents = await arquivo.read()
    if len(contents) <= 4:
        return err_json("Arquivo vazio ou inválido.")
    Path(ARQ_ENTRADA).write_bytes(contents)
    # guarda o hash do arquivo original pra dar XP uma única vez por rota (anti-fraude)
    sess["_rota_hash"] = gamification.calcular_hash_rota(contents)
    print(f"  [UPLOAD] {ARQ_ENTRADA} salvo ({len(contents)} bytes) — usuário: {sess['usuario']}")
    return ok_json({"ok": True})

@app.post("/scan/ocr")
async def scan_ocr(request: Request):
    """
    Recebe uma imagem em base64 (foto de etiqueta/pacote) e retorna o texto
    reconhecido via Google Cloud Vision API (OCR). Requer apenas login básico
    (não usa o gate de acesso pago do /upload, pois é uma feature separada).
    """
    _sessao_ou_401(request)

    if not GOOGLE_VISION_API_KEY:
        return err_json("OCR não configurado no servidor (GOOGLE_VISION_API_KEY ausente).", 500)

    data = await request.json()
    imagem_b64 = (data.get("imagem") or "").strip()
    if not imagem_b64:
        return err_json("Nenhuma imagem enviada.")

    # Remove prefixo data URL, se vier (ex.: "data:image/jpeg;base64,...")
    if "," in imagem_b64 and imagem_b64.strip().startswith("data:"):
        imagem_b64 = imagem_b64.split(",", 1)[1]

    payload = {
        "requests": [{
            "image": {"content": imagem_b64},
            "features": [{"type": "TEXT_DETECTION"}],
            "imageContext": {"languageHints": ["pt"]},
        }]
    }

    try:
        resp = requests.post(
            GOOGLE_VISION_URL,
            params={"key": GOOGLE_VISION_API_KEY},
            json=payload,
            timeout=30,
        )
        resp.raise_for_status()
        result = resp.json()
    except requests.exceptions.RequestException as e:
        print(f"  [SCAN/OCR] ❌ falha na chamada à Vision API: {e}")
        return err_json("Falha ao consultar o serviço de OCR. Tente novamente.", 502)

    resposta = (result.get("responses") or [{}])[0]
    if "error" in resposta:
        msg = resposta["error"].get("message", "Erro desconhecido na Vision API.")
        print(f"  [SCAN/OCR] ❌ Vision API: {msg}")
        return err_json(msg, 502)

    texto = ""
    annotations = resposta.get("textAnnotations") or []
    if annotations:
        texto = annotations[0].get("description", "")

    return ok_json({"ok": True, "texto": texto})

# ─── Anjun: CSV -> XLSX geocodificado, com progresso em tempo real ──
_ANJUN_JOBS: dict = {}  # uid -> {status, total, done, erro, pronto, nome_saida}

def _anjun_processar_bg(uid: str, entrada_path: Path, saida_path: Path):
    job = _ANJUN_JOBS[uid]
    try:
        env_here = {**os.environ, "HERE_API_KEY": HERE_API_KEY}
        proc = subprocess.Popen(
            [sys.executable, ANJUN_SCRIPT, str(entrada_path), str(saida_path)],
            stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
            text=True, bufsize=1, env=env_here,
        )
        for linha in proc.stdout:
            linha = linha.rstrip("\n")
            if linha:
                print(f"  [ANJUN] {linha}")
            m_total = re.search(r"Geocodificando (\d+) CEP", linha)
            if m_total:
                job["total"] = int(m_total.group(1))
            m_item = re.match(r"\s*\[(\d+)/(\d+)\]", linha)
            if m_item:
                job["done"] = int(m_item.group(1))
                job["total"] = int(m_item.group(2))
        proc.wait(timeout=600)
        if proc.returncode != 0:
            job["erro"] = "Erro ao rodar o script (verifique o CSV de entrada)."
        elif not saida_path.exists():
            job["erro"] = "O script rodou mas não gerou o arquivo de saída."
        else:
            job["total"] = max(job["total"], job["done"], 1)
            job["done"] = job["total"]
            job["pronto"] = True
    except subprocess.TimeoutExpired:
        job["erro"] = "Timeout: a geocodificação demorou mais que o esperado."
    except Exception as e:
        job["erro"] = str(e)
    finally:
        job["status"] = "concluido"

@app.post("/anjun/iniciar")
async def anjun_iniciar(request: Request, arquivo: UploadFile = File(...)):
    """
    Recebe um CSV (formato delivery_list, sem sequência), salva e dispara o
    csv_para_rota_xlsx.py em segundo plano. Requer só login básico — é uma
    ferramenta separada do pipeline principal, não consome crédito de
    importação. O progresso é consultado via GET /anjun/progresso.
    """
    sess = _sessao_ou_401(request)

    if not Path(ANJUN_SCRIPT).exists():
        return err_json(f"{ANJUN_SCRIPT} não encontrado na pasta do servidor.")

    contents = await arquivo.read()
    if len(contents) <= 4:
        return err_json("Arquivo vazio ou inválido.")

    uid = sess["user_id"]
    sufixo_original = Path(arquivo.filename or "entrada.csv").suffix or ".csv"
    entrada_path = DATA_DIR / f"anjun_entrada_{uid}{sufixo_original}"
    saida_path   = DATA_DIR / f"anjun_saida_{uid}.xlsx"
    entrada_path.write_bytes(contents)

    nome_saida = f"{Path(arquivo.filename or 'anjun').stem}_GEO.xlsx"
    _ANJUN_JOBS[uid] = {
        "status": "rodando", "total": 0, "done": 0,
        "erro": None, "pronto": False, "nome_saida": nome_saida,
        "saida_path": str(saida_path),
    }

    print(f"  [ANJUN] {entrada_path.name} salvo ({len(contents)} bytes) — usuário: {sess['usuario']}")

    thread = threading.Thread(target=_anjun_processar_bg, args=(uid, entrada_path, saida_path), daemon=True)
    thread.start()

    return ok_json({"ok": True})

@app.get("/anjun/progresso")
async def anjun_progresso(request: Request):
    sess = _sessao_ou_401(request)
    job = _ANJUN_JOBS.get(sess["user_id"])
    if not job:
        return err_json("Nenhum processamento em andamento.")
    return ok_json({
        "ok": True,
        "status": job["status"],
        "total": job["total"],
        "done": job["done"],
        "pronto": job["pronto"],
        "erro": job["erro"],
    })

@app.get("/anjun/baixar")
async def anjun_baixar(request: Request):
    sess = _sessao_ou_401(request)
    job = _ANJUN_JOBS.get(sess["user_id"])
    if not job or not job.get("pronto"):
        return err_json("Arquivo ainda não está pronto.")
    saida_path = Path(job["saida_path"])
    if not saida_path.exists():
        return err_json("Arquivo de saída não encontrado (pode ter expirado).")
    return FileResponse(
        str(saida_path),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=job["nome_saida"],
    )

@app.post("/pipeline")
async def pipeline(request: Request):
    sess = _sessao_com_acesso_ou_403(request)
    if not Path(ARQ_ENTRADA).exists():
        return err_json(f"{ARQ_ENTRADA} não encontrado. Faça o upload primeiro.")
    if not Path(TRATAMENTO_PY).exists():
        return err_json(f"{TRATAMENTO_PY} não encontrado na pasta.")
    print(f"\n  [PIPELINE] Rodando {TRATAMENTO_PY}...")
    try:
        env_here = {**os.environ, "HERE_API_KEY": HERE_API_KEY, "HERE_CIDADE_UF": HERE_CIDADE_UF}
        result = subprocess.run(
            [sys.executable, TRATAMENTO_PY],
            capture_output=True, text=True, timeout=600, env=env_here
        )
        if result.returncode != 0:
            erro = result.stderr or result.stdout or "Erro desconhecido"
            print(f"  [PIPELINE] ❌ {erro}")
            return err_json(erro)
        print(f"  [PIPELINE] ✅ Pipeline concluído")
        if result.stdout:
            print(result.stdout)
        rows, headers = ler_processado(sess["user_id"])
        sess["dados"] = (rows, headers)
        arq_final = ARQ_VALIDADO if Path(ARQ_VALIDADO).exists() else ARQ_PROCESSADO
        adicionar_ao_historico(Path(arq_final).name, rows, headers, sess["user_id"])
        if not sess.get("is_admin"):
            usuario_consumir_credito_avulso_se_necessario(sess["usuario"])
            registrar_importacao_hoje(sess["usuario"])
        print(f"  [PIPELINE] ✅ {len(rows)} endereços carregados")

        xp_resultado = None
        rota_hash = sess.get("_rota_hash")
        if rota_hash:
            paradas = len(rows)
            pacotes = sum(int(r.get("group_size") or 1) for r in rows)
            xp_resultado = _conceder_xp_rota(sess["user_id"], paradas, pacotes, rota_hash)
            sess["_rota_hash"] = None  # cada upload só pode gerar XP uma vez

        resposta = {"ok": True, "total": len(rows)}
        if xp_resultado:
            resposta["gamificacao"] = xp_resultado
        return ok_json(resposta)
    except subprocess.TimeoutExpired:
        return err_json("Timeout: o pipeline demorou mais que o esperado.")
    except Exception as e:
        print(f"  [PIPELINE] ❌ {e}")
        return err_json(str(e))

@app.post("/rota/otimizar")
async def rota_otimizar(request: Request):
    """
    Reordena os endereços já carregados na sessão pela sequência mais
    eficiente de entrega, calculada via OSRM (dados OpenStreetMap).
    Endereços sem coordenada válida são mantidos, mas empurrados pro
    final da lista (não entram no cálculo de otimização).
    """
    sess = _sessao_ou_401(request)
    if sess["dados"] is None:
        return err_json("Nenhum dado processado ainda.", 404)
    rows, headers = sess["dados"]

    com_coord = []
    sem_coord = []
    for row in rows:
        try:
            lat = float(row.get("lat") or "")
            lon = float(row.get("lon") or "")
            com_coord.append((row, lat, lon))
        except (TypeError, ValueError):
            sem_coord.append(row)

    if len(com_coord) < 2:
        return err_json("É necessário pelo menos 2 endereços com coordenadas válidas para otimizar a rota.")

    coords = [(lat, lon) for _, lat, lon in com_coord]
    ordem, servico = _otimizar_sequencia(coords)
    if ordem is None:
        return err_json(
            "Não foi possível calcular a rota otimizada agora (serviços de roteamento indisponíveis). "
            "Tente novamente em instantes.", 502
        )

    rows_otimizadas = [com_coord[i][0] for i in ordem] + sem_coord
    sess["dados"] = (rows_otimizadas, headers)

    print(f"  [OTIMIZAR] {sess['usuario']} otimizou {len(com_coord)} paradas via {servico} "
          f"({len(sem_coord)} sem coordenada ficaram no final)")

    return ok_json({
        "ok": True,
        "rows": rows_otimizadas,
        "headers": headers,
        "sem_coordenadas": len(sem_coord),
        "servico": servico,
    })

# ═══════════════════════════════════════════════════════════════════
#  LOTES DE TERCEIROS (Aparecida de Goiânia / Senador Canedo / Goiânia)
#  ──────────────────────────────────────────────────────────────────
#  Proxy servidor-a-servidor para os vector tiles (.pbf) de quadra/lote
#  usados no Route Planner. Evita problema de CORS no navegador, esconde
#  o token do usuário final, e usa um cache SQLite PERMANENTE (em
#  DATA_DIR, mesmo Volume do Railway usado pro resto do app) — depois do
#  primeiro request (ou de rodar warmup_lotes.py), o tile nunca mais
#  precisa ser buscado de novo no routeplanner.com.br.
#  Lógica de cache/fetch fica em lotes_terceiros_cache.py, compartilhada
#  com o script de warm-up (warmup_lotes.py).
# ═══════════════════════════════════════════════════════════════════

from starlette.concurrency import run_in_threadpool
from lotes_terceiros_cache import get_tile as _lotes_get_tile
from lotes_terceiros_cache import LOTES_TERCEIROS_CIDADES as _LOTES_CIDADES


@app.get("/api/lotes-terceiros/{cidade}/{z}/{x}/{y}.pbf")
async def lotes_terceiros_tile(cidade: str, z: int, x: int, y: int):
    """Repassa um vector tile (.pbf) de quadra/lote de Aparecida de Goiânia,
    Senador Canedo ou Goiânia — lendo do cache permanente sempre que possível
    e só caindo pro Route Planner se o tile nunca foi visto antes."""
    if cidade not in _LOTES_CIDADES:
        raise HTTPException(status_code=400, detail="Cidade inválida (use 'aparecida', 'canedo' ou 'goiania').")

    try:
        data = await run_in_threadpool(_lotes_get_tile, cidade, z, x, y)
    except Exception as e:
        print(f"  [LOTES-TERCEIROS] falha ao buscar tile {cidade} {z}/{x}/{y}: {type(e).__name__}: {e}")
        raise HTTPException(status_code=502, detail="Falha ao buscar tile.")

    if data is None:
        # Tile sem lote cadastrado nessa área — normal, não é erro (e já
        # ficou salvo no cache como tombstone, não busca de novo).
        return Response(status_code=204)

    return Response(
        content=data,
        media_type="application/x-protobuf",
        headers={"Cache-Control": "public, max-age=604800"},  # cache 7 dias no navegador
    )


# ═══════════════════════════════════════════════════════════════════
#  ROTAS  ──  DELETE
# ═══════════════════════════════════════════════════════════════════

@app.delete("/admin/usuarios")
async def admin_apagar_usuario_route(request: Request, usuario: str = ""):
    sess = _sessao_admin_ou_403(request)
    ok, msg = admin_apagar_usuario(usuario, sess["usuario"])
    return ok_json({"ok": ok, "msg": msg})

@app.delete("/historico/apagar")
async def historico_apagar(request: Request, nome: str = ""):
    sess = _sessao_ou_401(request)
    historico = carregar_historico()
    existia = any(h.get("nome") == nome and h.get("user_id") == sess["user_id"] for h in historico)
    novo = [h for h in historico if not (h.get("nome") == nome and h.get("user_id") == sess["user_id"])]
    salvar_historico(novo)
    if existia:
        # Rota saiu do histórico dele — as correções manuais que ele fez
        # nela eram só temporárias, então voltam ao valor global normal.
        banco_coords_limpar_overrides_usuario(sess["user_id"])
    return ok_json({"ok": True})

# ═══════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════

# Registrado por último de propósito — ver comentário lá em cima, perto
# do app.mount de /static/icons.
app.include_router(gamification.router)

if __name__ == "__main__":
    print(f"""
╔══════════════════════════════════════════════════╗
║       ROTA MANAGER — SERVIDOR  (FastAPI)         ║
╠══════════════════════════════════════════════════╣
║  Endereço : http://{HOST}:{PORT}
║  Pasta    : {Path('.').resolve()}
╚══════════════════════════════════════════════════╝
""")
    try:
        import openpyxl  # noqa
    except ImportError:
        print("⚠️  openpyxl não encontrado. Instalando...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])

    _migrar_para_volume()
    _bootstrap_admin()

    uvicorn.run(
        app,
        host=HOST,
        port=PORT,
        workers=1,          # múltiplos workers incompatíveis com sessões em memória
        log_level="info",
    )
